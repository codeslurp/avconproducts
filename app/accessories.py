"""Accessories loader.

Accessories are intentionally NOT a cascade-resolution catalog like valves and
actuators — the source file is a consolidated dashboard where 14 different
product families are stacked into one sheet, each with its OWN column schema
embedded as a header row. There's no single set of cascade dropdowns that
makes sense across all families, so the UI just browses the list with
checkboxes and lets the user pick any combination.

This module produces a flat list of `Accessory` dicts ready for the API to
serve. It filters out two flavors of garbage we know exist in the source:

  1. Header rows: each family's first row repeats the column headers
     (`code='Code'`, `model='Model'`). These appear as fake products if
     loaded blindly.
  2. Stray `Table_*` rows: 16 ball-valve rows leaked into the file. They
     have a family name like 'Table_2090F BV NEW CODEING' and no Code.

When/if engineering cleans up the source file, the filters here become
no-ops and can be removed — but they're cheap to keep and document.
"""
from __future__ import annotations

import shutil
import tempfile
import warnings
from collections import OrderedDict
from pathlib import Path
from typing import Any

import openpyxl


SHEET_NAME = "Accessories"
# Source column 1 holds the family marker (e.g. "ALR Data", "MOR", "BKT").
FAMILY_COL = 1
CODE_COL = 2


def _norm(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "")
        return s if s else None
    return v


def _clean_family(raw: str) -> str:
    """Normalize family names — strip the redundant ' Data' suffix some have
    (e.g. 'ALR Data' -> 'ALR', 'FCV Data' -> 'FCV'), keep others as-is."""
    s = str(raw).strip()
    if s.endswith(" Data"):
        s = s[: -len(" Data")]
    return s


# Single-family accessory files with their own tabular layout (code in a key
# column, row 1 = headers). Each becomes ONE accessory family appended to the
# consolidated list. Added 2026-06-02: Positioner + Solenoid Valve.
EXTRA_ACCESSORY_SOURCES = [
    # Two valve-positioner workbooks, renamed 2026-06-04 to unambiguous names:
    #   "Electronic Valve Positioner.xlsx" (ED codes)   -> family "EVP"
    #   "Pneumatic Valve Positioner.xlsx"  (PS/PD codes) -> family "Positioner" (shown as EPP)
    # The old "…Positioner Data Sheet…" name collision and its exclude trick are gone.
    {"file_substring": "Electronic Valve Positioner", "sheet": "Positioner", "family": "EVP",            "code_col": 1},
    {"file_substring": "Pneumatic Valve Positioner",  "sheet": "Positioner", "family": "Positioner",     "code_col": 1},
    {"file_substring": "Solenoid Valve",  "sheet": "Sheet1",     "family": "Solenoid Valve", "code_col": 1},
    # THW FOR MSD now has its own dedicated file (cleaner columns: Suitable
    # Actuator Model, Material, …). Loaded here and SKIPPED from the consolidated
    # Dashboard sheet (see `dedicated_families` in load_accessories) so it isn't
    # loaded twice. The R1_Updated export renamed the sheet "THW FOR MSD" ->
    # "Sheet1"; both are listed so whichever THW file wins the most-recent pick
    # still loads.
    {"file_substring": "THW FOR MSD",     "sheet": ["Sheet1", "THW FOR MSD"], "family": "THW FOR MSD", "code_col": 1},
    # Migrated 2026-06-07 from the consolidated Dashboard sheet to dedicated
    # "New Structure" single-family files (cleaner per-family column schemas).
    # Each is skipped from the consolidated load via `dedicated_families`.
    {"file_substring": "Gland Data",      "sheet": "Gland Data",   "family": "Gland",   "code_col": 1},
    # merge_all: the two MOR files hold DISJOINT vendor ranges, not successive
    # revisions — R1 = 13 Q-Tork (MRQ*) + 10 Transtork (MRT*), R1_Updated = 13
    # Shakti (MRS*), zero overlap. Newest-wins would silently drop 23 SKUs.
    {"file_substring": "Manual Override", "sheet": "MOR",          "family": "MOR",     "code_col": 1, "merge_all": True},
    {"file_substring": "Plug Data",       "sheet": "Plug Data",    "family": "Plug",    "code_col": 1},
    # Tube & Fittings, migrated 2026-08-02 — the deferral note said "source row
    # count under review"; cell-diffed and resolved: the dedicated file and the
    # consolidated sheet hold the SAME 6 codes with the SAME values. The only
    # difference is the label, and the dedicated file has it right — the
    # consolidated sheet's shared header row calls the material "Model"
    # (e.g. "Model = SS304"), where it is a Material of Construction.
    # file_substring is "Fitting Data", NOT "Tube Fitting": the filename has a
    # DOUBLE space ("Tube  Fitting Data for New Structure_R1.xlsx"), so a
    # single-spaced substring would silently fail to match and the family would
    # quietly fall back to the consolidated sheet.
    {"file_substring": "Fitting Data",    "sheet": "FITTING Data", "family": "FITTING", "code_col": 1},
]


# Curated, ORDERED accessory families. The picker shows families in THIS order;
# any loaded family not listed here is appended after (e.g. "THW FOR MSD").
# Each entry: (data_key, tag, letter, display_name)
#   data_key  = the family string as it appears in the loaded data (the cleaned
#               consolidated-sheet marker, or an EXTRA_ACCESSORY_SOURCES family).
#               None = no data yet → rendered as a "data pending" placeholder.
#   tag       = short chip label shown in the UI.
#   letter    = single-char code used in the COMBINED PRODUCT CODE (must be
#               unique across families; e.g. Silencer uses 'Z' because 'S' is
#               taken by Solenoid Valve).
#   display_name = full human name.
# NOTE: data_key "Positioner" is the PNEUMATIC valve positioner (PS/PD codes),
# shown as EPP — "Electronic Pneumatic Positioner". data_key "EVP" (ED codes) is
# shown as ECC — "Electronic Control Card". Renamed 2026-06-11 from PVP/EVP (tag,
# letter, and display_name only); data_keys are kept stable so the data load and
# the actuator-gating rulesets (POSITIONER_KEY / allow-lists in app.js, which key
# on data_keys) still match. EPP took letter 'P'; Plug moved 'P' -> 'U'.
ACCESSORY_FAMILY_ORDER: list[tuple[str | None, str, str, str]] = [
    ("Solenoid Valve",  "SV",                  "S", "Solenoid Valve"),
    ("LSB",             "LSB",                 "L", "Limit Switch"),
    ("EVP",             "ECC",                 "E", "Electronic Control Card"),
    ("Positioner",      "EPP",                 "P", "Electronic Pneumatic Positioner"),
    ("MOR",             "MOR",                 "M", "Manual Override"),
    ("FRG",             "AFR",                 "R", "Air Filter Regulator"),
    ("CFLG",            "CFLG",                "C", "Companion Flange"),
    ("Gland",           "Gland",               "G", "Gland"),
    ("Plug",            "Plug",                "U", "Plug"),
    ("Silencer",        "Silencer/Bug Screen", "Z", "Silencer/Bug Screen"),
    ("Volume Booster",  "Volume Booster",      "V", "Volume Booster"),
    ("FITTING",         "Tube & Fittings",     "N", "Tube & Fittings"),
    ("FCV",             "FCV",                 "F", "Flow Control Valve"),
    ("ALR",             "ALR",                 "A", "Air Lock Relay"),
    ("QEV",             "QEV",                 "Q", "Quick Exhaust Valve"),
    (None,              "Direct Mount",        "D", "Direct Mount"),
    ("BKT",             "BKT",                 "B", "Bracket Mount"),
    ("THW FOR MSD",     "THW",                 "H", "THW for MSD"),  # data_key (elem 0) must match the loaded family; display name = "THW for MSD"
]


# "Flag" families have no SKU catalog by nature — they're a yes/no option (e.g.
# a mounting style). They are ADDABLE (contribute their letter to the combined
# product code) but carry no data, so they must NOT render as "data pending".
# Keyed by family key (= data_key, or the tag for data_key=None entries).
FLAG_FAMILIES = {"Direct Mount"}


def _meta_by_data_key() -> dict[str, tuple[str, str, str]]:
    """data_key -> (tag, letter, display_name) for families that have data."""
    return {
        dk: (tag, letter, label)
        for dk, tag, letter, label in ACCESSORY_FAMILY_ORDER
        if dk is not None
    }


def _build_ordered_families(counts: dict[str, int]) -> list[dict[str, Any]]:
    """Build the picker's family list in curated order. `counts` is data_key ->
    row count (from the loaded rows). Declared-but-empty families (data_key None,
    or present in the order but with zero rows) are marked pending=True. Any
    loaded family NOT in the curated order is appended last (e.g. THW FOR MSD)."""
    used: set[str] = set()
    families: list[dict[str, Any]] = []
    for data_key, tag, letter, label in ACCESSORY_FAMILY_ORDER:
        count = counts.get(data_key, 0) if data_key else 0
        key = data_key if data_key else tag
        is_flag = key in FLAG_FAMILIES
        families.append({
            "key": key,
            "tag": tag,
            "letter": letter,
            "label": label,
            "count": count,
            # Flag families are addable-but-codeless: never "pending".
            "pending": (not is_flag) and ((data_key is None) or count == 0),
            "flag": is_flag,
        })
        if data_key:
            used.add(data_key)
    for data_key, count in counts.items():
        if data_key in used:
            continue
        families.append({
            "key": data_key, "tag": data_key, "letter": "",
            "label": data_key, "count": count, "pending": False, "flag": False,
        })
    return families


def _attach_family_meta_to_rows(rows: list[dict[str, Any]]) -> None:
    """Stamp each row with its family's `tag` (chip label) and `letter` (product-
    code char). Families not in the curated order fall back to tag=family name,
    letter='' (so the combined code falls back to the full accessory code)."""
    meta = _meta_by_data_key()
    for r in rows:
        tag, letter, _label = meta.get(r["family"], (r["family"], "", r["family"]))
        r["tag"] = tag
        r["letter"] = letter


def find_accessories_file(data_dir: Path) -> Path | None:
    """Pick the consolidated accessories workbook — the most recently modified
    .xlsx in data/Accessories/ that actually contains an `Accessories` sheet.
    Checking for the sheet (not just newest file) stops single-family files like
    'Solenoid Valve as accessories.xlsx' from hijacking the load. Returns None if
    the folder is absent/empty — the app stays usable without accessories."""
    acc_dir = data_dir / "Accessories"
    if not acc_dir.exists():
        return None
    candidates = [
        p for p in acc_dir.rglob("*.xls*")
        if not p.name.startswith("~$") and p.name not in _extra_filenames(acc_dir)
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates:
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb = openpyxl.load_workbook(p, read_only=True, keep_vba=False)
            has_sheet = SHEET_NAME in wb.sheetnames
            wb.close()
            if has_sheet:
                return p
        except Exception:
            continue
    return candidates[0]  # fallback: nothing had an Accessories sheet


def _extra_filenames(acc_dir: Path) -> set[str]:
    """Filenames that are single-family extra sources (so the consolidated-file
    search skips them)."""
    names: set[str] = set()
    for src in EXTRA_ACCESSORY_SOURCES:
        for p in acc_dir.rglob("*.xls*"):
            if src["file_substring"] in p.name and not p.name.startswith("~$"):
                names.add(p.name)
    return names


def _load_extra_family(acc_dir: Path, src: dict) -> list[dict[str, Any]]:
    """Load one accessory family from its dedicated file(s).

    Default: the most-recently-modified match wins — a later export supersedes
    an earlier one for the same family.

    With `merge_all: True`: EVERY match is read, oldest first, and rows are
    merged on `code`. Used where separate files hold DISJOINT vendor ranges of
    one family rather than successive revisions of the same range — MOR ships
    as Q-Tork + Transtork (R1, 23 rows) plus Shakti (R1_Updated, 13 rows) with
    zero code overlap. Newest-wins on a genuine duplicate code preserves
    supersede semantics, and every override is logged rather than silent.
    """
    cands = [
        p for p in acc_dir.rglob("*.xls*")
        if src["file_substring"] in p.name and not p.name.startswith("~$")
        and (not src.get("exclude") or src["exclude"] not in p.name)
    ]
    if not cands:
        return []
    if not src.get("merge_all"):
        newest = sorted(cands, key=lambda p: p.stat().st_mtime, reverse=True)[0]
        return _read_family_file(newest, src)

    merged: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    for path in sorted(cands, key=lambda p: p.stat().st_mtime):
        for row in _read_family_file(path, src):
            if row["code"] in merged:
                print(
                    f"[valve-selector] Accessories: {src['family']} code "
                    f"{row['code']} overridden by newer {path.name}.",
                    flush=True,
                )
            merged[row["code"]] = row
    print(
        f"[valve-selector] Accessories: merged {len(merged)} {src['family']} "
        f"rows from {len(cands)} file(s).",
        flush=True,
    )
    return list(merged.values())


def _read_family_file(path: Path, src: dict) -> list[dict[str, Any]]:
    """Read ONE single-family accessory workbook into [{code, family, attrs}]."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False, read_only=True)
        except PermissionError:
            tmp_dir = Path(tempfile.gettempdir()) / "valve-selector-cache"
            tmp_dir.mkdir(exist_ok=True)
            tmp_path = tmp_dir / path.name
            shutil.copyfile(path, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, data_only=True, keep_vba=False, read_only=True)
    # `sheet` may be a single name or a list of candidate names (sources whose
    # sheet got renamed across exports list both — e.g. THW's R1_Updated renamed
    # "THW FOR MSD" -> "Sheet1"). Use the first candidate present in the file.
    wanted = src["sheet"] if isinstance(src["sheet"], (list, tuple)) else [src["sheet"]]
    sheet_name = next((s for s in wanted if s in wb.sheetnames), None)
    if sheet_name is None:
        wb.close()
        print(f"[valve-selector] Accessories: none of {wanted} found in {path.name}; skipping.", flush=True)
        return []
    raw_rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    if not raw_rows:
        return []
    headers = [str(_norm(h) or "").strip() for h in raw_rows[0]]
    code_idx = src["code_col"] - 1
    out: list[dict[str, Any]] = []
    for raw in raw_rows[1:]:
        if not raw or code_idx >= len(raw):
            continue
        code = _norm(raw[code_idx])
        if code in (None, ""):
            continue
        code = str(code).strip()
        if code_idx < len(headers) and code.lower() == headers[code_idx].lower():
            continue  # repeated header row
        attrs = []
        for i, val in enumerate(raw):
            if i == code_idx:
                continue
            v = _norm(val)
            if v in (None, ""):
                continue
            label = headers[i] if i < len(headers) else f"Col {i + 1}"
            attrs.append({"label": label, "value": str(v)})
        out.append({"code": code, "family": src["family"], "attrs": attrs})
    print(f"[valve-selector] Accessories: loaded {len(out)} {src['family']} rows from {path.name}.", flush=True)
    return out


def load_accessories(data_dir: Path) -> dict[str, Any]:
    """Returns a dict with `rows`, `families`, `headers`, `source_file`.
    Rows are filtered (no header-row pollution, no Table_* garbage).
    `headers` is the row-1 column names from the source — used by the UI
    to label attribute values in the detail view."""
    src = find_accessories_file(data_dir)
    if src is None:
        return {"rows": [], "families": [], "headers": [], "source_file": None}

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb = openpyxl.load_workbook(
                src, data_only=True, keep_vba=False, read_only=True
            )
        except PermissionError:
            # Source open in Excel; copy to temp and read that.
            tmp_dir = Path(tempfile.gettempdir()) / "valve-selector-cache"
            tmp_dir.mkdir(exist_ok=True)
            tmp_path = tmp_dir / src.name
            shutil.copyfile(src, tmp_path)
            wb = openpyxl.load_workbook(
                tmp_path, data_only=True, keep_vba=False, read_only=True
            )

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        print(
            f"[valve-selector] Accessories: no '{SHEET_NAME}' sheet in {src.name}; "
            f"skipping.",
            flush=True,
        )
        return {"rows": [], "families": [], "headers": [], "source_file": src.name}

    ws = wb[SHEET_NAME]
    raw_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not raw_rows:
        return {"rows": [], "families": [], "headers": [], "source_file": src.name}

    headers = [str(_norm(h) or "").strip() for h in raw_rows[0]]
    # The header for col 1 in the source is itself a family tag ("ALR Data"),
    # not a real header — replace with "Family" so the UI can label it.
    if headers and headers[0]:
        headers[0] = "Family"

    # PASS 1: collect each family's OWN header row.
    #
    # The sheet stacks ~14 families, each preceded by a header row (its Code
    # column literally reads "Code"). Those rows used to be discarded as garbage
    # and every family was then labelled with row 1 — which is just ALR's schema.
    # That mislabelled every other family, with a different effective offset each
    # (LSB showed "Make = SB202L2" for what is a Model of LSB; Volume Booster
    # showed "Make = SS316" for a Material of Construction; and so on).
    #
    # Two passes rather than one so a family whose data precedes its header row
    # would still be labelled correctly. Row 1 doubles as ALR's header, so it is
    # registered here too. Values are untouched — labels only.
    family_headers: dict[str, list[str]] = {}

    def _register_header(raw_row) -> None:
        fam_raw = _norm(raw_row[FAMILY_COL - 1])
        if not fam_raw:
            return
        family_headers[_clean_family(fam_raw)] = [
            str(_norm(h) or "").strip() for h in raw_row
        ]

    _register_header(raw_rows[0])
    for raw in raw_rows[1:]:
        if not raw:
            continue
        if str(_norm(raw[CODE_COL - 1]) or "").strip().lower() == "code":
            _register_header(raw)

    # PASS 2: build the rows.
    rows: list[dict[str, Any]] = []
    families_seen: "OrderedDict[str, int]" = OrderedDict()
    skipped_header = 0
    skipped_table = 0
    skipped_no_code = 0
    skipped_dedicated = 0
    # Families that now load from a dedicated extra-source file (e.g. THW FOR MSD)
    # are skipped here, so they aren't loaded from BOTH the consolidated sheet and
    # the dedicated file.
    dedicated_families = {s["family"] for s in EXTRA_ACCESSORY_SOURCES}

    for raw in raw_rows[1:]:
        if not raw or all(v is None for v in raw):
            continue
        family_raw = _norm(raw[FAMILY_COL - 1])
        code_raw = _norm(raw[CODE_COL - 1])

        # FILTER 1: stray Table_* leak from another catalog
        if family_raw and str(family_raw).startswith("Table_"):
            skipped_table += 1
            continue
        # FILTER 2: a family's own embedded header row (e.g. row with code='Code')
        if str(code_raw or "").strip().lower() == "code":
            skipped_header += 1
            continue
        # FILTER 3: rows with no code aren't useful for selection
        if not code_raw:
            skipped_no_code += 1
            continue

        family = _clean_family(family_raw) if family_raw else "Other"
        # FILTER 4: family superseded by a dedicated extra-source file (loaded
        # separately below) — skip the consolidated copy to avoid duplicates.
        if family in dedicated_families:
            skipped_dedicated += 1
            continue
        # Build attributes dict — pair headers with values, drop empty cells.
        # Labels come from THIS family's own header row (see pass 1); the sheet-
        # wide row-1 headers are only a fallback for a family that somehow has
        # none of its own.
        fam_hdrs = family_headers.get(family, headers)
        attrs = []
        for i, val in enumerate(raw):
            v = _norm(val)
            if i == FAMILY_COL - 1 or i == CODE_COL - 1:
                continue  # family + code are top-level fields
            if v is None or v == "":
                continue
            label = fam_hdrs[i] if i < len(fam_hdrs) and fam_hdrs[i] else (
                headers[i] if i < len(headers) else f"Col {i + 1}"
            )
            attrs.append({"label": label, "value": str(v)})

        rows.append({
            "code": str(code_raw).strip(),
            "family": family,
            "attrs": attrs,
        })
        families_seen[family] = families_seen.get(family, 0) + 1

    print(
        f"[valve-selector] Accessories: loaded {len(rows)} rows across "
        f"{len(families_seen)} families from {src.name} "
        f"(skipped: {skipped_header} headers, {skipped_table} Table_* "
        f"leaks, {skipped_no_code} no-code rows, {skipped_dedicated} dedicated-file dupes).",
        flush=True,
    )

    # Append single-family extra sources (Positioner→EPP, Solenoid Valve→SV) —
    # each its own file/family, merged into the same flat list the UI browses.
    acc_dir = data_dir / "Accessories"
    for esrc in EXTRA_ACCESSORY_SOURCES:
        extra = _load_extra_family(acc_dir, esrc)
        if extra:
            rows.extend(extra)

    # Order + label families per the curated ACCESSORY_FAMILY_ORDER, attach the
    # chip tag + product-code letter to every row, and inject declared-but-empty
    # families (EVP, Plug, Direct Mount) as 'data pending' placeholders.
    counts: dict[str, int] = {}
    for r in rows:
        counts[r["family"]] = counts.get(r["family"], 0) + 1
    families = _build_ordered_families(counts)
    _attach_family_meta_to_rows(rows)

    return {
        "rows": rows,
        "families": families,
        "headers": headers,
        "source_file": src.name,
    }
