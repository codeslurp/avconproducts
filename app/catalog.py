"""In-memory product catalogs (valves and actuators).

A `ValveTypeConfig` describes one product family — its cascade order, detail
columns, which file to read, and which sheets in that file are catalogs.

`load_all` reads every registered type at startup; types whose source file
isn't present in `data/` are skipped silently so the app still launches with
a subset.
"""
from __future__ import annotations

import shutil
import tempfile
import warnings
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class PairedActuator:
    """One catalog-recommended actuator for a valve resolution.

    A valve row may name several recommended actuators in different columns
    (e.g. col 49 = pneumatic model, col 100 = electric rotary model).
    Each gets its own card in the result UI, pointing into the target
    catalog's picker with `target_field` pre-set to `model`.

    Use `target_type` for a single-catalog routing (e.g. all electric
    recommendations go to `electrical_rotary`).

    Use `target_type_by_prefix` when one recommendation column can name
    actuators from MULTIPLE catalogs and the right catalog is determined by a
    prefix on the model name. For example, butterfly valves' pneumatic column
    mixes `ACT-*` (Rack & Pinion) and `SYA-*` (Scotch Yoke); we route based
    on which prefix the recommended model starts with. The tuples are checked
    in order; first matching prefix wins. If no prefix matches, `target_type`
    in the API response is `None` and the front-end should show the model as
    informational text without a "jump to catalog" button.
    """
    model_col: int               # 1-based col on the valve row that names the model
    target_field: str            # cascade key in target to pre-fill (e.g. "model")
    label: str                   # heading shown above the card
    target_type: str | None = None
    target_type_by_prefix: tuple[tuple[str, str], ...] = ()
    # When set, a recommendation cell whose (normalized) value does NOT start
    # with this prefix is ignored entirely — no card. Used for dual-use columns
    # that mix actuator model codes with other content (e.g. the control valve
    # "Fail Safe" columns hold MSD-* models OR a pressure/0). Default None keeps
    # the historical behaviour (every non-empty cell becomes a card).
    require_prefix: str | None = None
    # Optional per-series card-label overrides. Maps a Series-value PREFIX
    # (e.g. "5016A") to the label to show instead of `label`. Used so 5016
    # shows "A Port Close"/"B Port Close" while other series keep the generic
    # Fail-Safe labels. Prefix match against the row's Series (col 5).
    series_labels: dict = field(default_factory=dict)

    def label_for(self, series_value: str | None) -> str:
        """Per-series card label: first matching prefix in series_labels, else label."""
        if series_value:
            sv = str(series_value).strip()
            for prefix, lab in self.series_labels.items():
                if sv.startswith(prefix):
                    return lab
        return self.label

    def __post_init__(self) -> None:
        if bool(self.target_type) == bool(self.target_type_by_prefix):
            raise ValueError(
                f"PairedActuator must set exactly one of `target_type` or "
                f"`target_type_by_prefix` (got target_type={self.target_type!r}, "
                f"target_type_by_prefix={self.target_type_by_prefix!r})."
            )

    def resolve_target_type(self, model: str) -> str | None:
        """Pick the target catalog key for a recommended `model` name."""
        if self.target_type_by_prefix:
            for prefix, tt in self.target_type_by_prefix:
                if model.startswith(prefix):
                    return tt
            return None
        return self.target_type


@dataclass(frozen=True)
class EnrichmentSource:
    """VLOOKUP-style join: pulls columns from another workbook into the master
    catalog, keyed by a shared column (e.g. Bare Valve Code).

    Used by the master Ball Valve catalog to fold in the "Electric Actuator"
    columns that live only in the 4 per-series .xlsx extracts (2030F / 2060F /
    2070F / 2090F BV NEW CODEING)."""
    file_substring: str
    sheet_marker: str
    source_key_col: int                          # 1-based col in source = join key
    master_key_col: int                          # 1-based col in master = join key
    columns: tuple[tuple[int, int], ...]         # (source_col, master_dest_col) pairs


@dataclass(frozen=True)
class ExtraRowSource:
    """Adds ROWS (not columns) from a secondary workbook whose SKU keys aren't
    already in the catalog. Source columns are remapped to the catalog's column
    positions via `column_map`. Used to fold in SKUs that live only in a
    secondary file with a different layout (e.g. the 162 Pharma SKUs that exist
    only in 'Pharma Dashboard.xlsx', or actuators present only in a combined
    workbook). Non-destructive: nothing is written back to any source file.

    Dedup: a source row is added only if its key isn't already in the catalog,
    so re-listing existing SKUs is harmless. `key_pattern` (optional regex) gates
    which source rows count — used to skip embedded header/garbage rows."""
    file_substring: str
    sheet_marker: str
    key_col: int                                 # 1-based source col holding the SKU key
    master_key_col: int                          # 1-based catalog col the key maps to
    column_map: tuple[tuple[int, int], ...]      # (source_col, master_col) pairs
    path_contains: str | None = None
    key_pattern: str | None = None               # only add rows whose key matches


# Sentinel option offered for a "nullable" cascade field (see
# ValveTypeConfig.nullable_cascade_keys) when some matching rows have a blank in
# that field. Selecting it filters to the blank-valued rows — so a SKU with no
# value for that attribute (e.g. a control valve with no Certification) is
# reachable instead of being force-matched to the sole non-blank value. Must be
# a string no real datum equals; mirrored verbatim in docs/static/catalog-engine.js.
CASCADE_NULL_OPTION = "Not Applicable"


def _is_blank_cell(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


@dataclass(frozen=True)
class ValveTypeConfig:
    key: str                            # URL/JSON identifier, e.g. "ball"
    label: str                          # Section title shown in the UI
    category: str                       # Top-level picker (e.g. "Valves", "Actuators").
                                        # Each distinct category gets its own picker widget.
    file_substring: str                 # substring used to pick the file from data/
    sheet_marker: str                   # substring marking which sheets are catalogs
    cascade: list[tuple[str, int, str]] # ordered (form-key, 1-based col, label)
    detail_columns: list[tuple[int, str]]  # (1-based col, display label)

    # Sub-grouping shown as a heading inside the picker menu (None = no heading).
    # E.g. within Actuators picker: "Pneumatic" vs "Electrical".
    subgroup: str | None = None

    # Optional PARENT heading above `subgroup`, for a two-level menu nesting.
    # When set, the menu renders `group` once as a parent header, then each
    # distinct `subgroup` under it as an indented child header. E.g. Pneumatic
    # actuators set group="Pneumatic" with subgroup="Rotary"/"Linear" so the
    # menu reads Pneumatic › Rotary / Linear. None = no parent (flat subgroup).
    group: str | None = None

    # Optional path scope to disambiguate same-named files across subfolders.
    # find_catalog_file matches on FILENAME only, so when two folders hold a
    # file with the same name (e.g. Pune and Mumbai both ship a
    # "Valve_Code_Selector_Dashboard.xlsx"), set this to a path fragment that
    # only the intended file's path contains (e.g. "Mumbai"). None = no scoping.
    path_contains: str | None = None

    # Headline result cards. Defaults match valves. Actuators override these
    # because they have a Code + Model (not a separate Catalogue Code) and
    # no torque-vs-FOS calculation.
    primary_label: str = "Bare Valve Code"
    primary_col: int = 1
    secondary_label: str | None = "Catalogue Code"
    secondary_col: int | None = 2
    show_bto_fos: bool = True
    bto_col: int = 39                   # only used if show_bto_fos

    # Optional one-line notice rendered in the result panel, for types that ship
    # with data the source hasn't supplied yet. Used INSTEAD of blank cards: the
    # relevant columns are left out of detail_columns and this explains why.
    # Remove it (and restore the columns) when the source revision lands.
    pending_note: str | None = None

    # Optional pairings: each entry in the tuple becomes its own
    # "Catalog-recommended actuator" card in the result panel, with a button
    # that jumps into that catalog's picker. Ball valves declare TWO: a
    # pneumatic R&P model (col 49) and an electric rotary model (col 100,
    # enriched in from the per-series files).
    paired_actuators: tuple[PairedActuator, ...] = ()

    # Optional VLOOKUP-style enrichment sources merged in after the main load.
    enrichment_sources: tuple[EnrichmentSource, ...] = ()

    # Optional secondary files that ADD rows (SKUs missing from the main file).
    extra_row_sources: tuple[ExtraRowSource, ...] = ()

    # Optional per-series cascade field-visibility overrides. Maps a value of
    # the `cascade_override_key` field (matched by PREFIX) to the ordered list
    # of cascade keys that stay visible for that value. Fields not listed are
    # hidden and skipped in resolution. Empty = same fields for every series.
    cascade_overrides: dict = field(default_factory=dict)
    # Which cascade key's value selects the override; defaults to the first
    # cascade key (e.g. Series) when cascade_overrides is set.
    cascade_override_key: str | None = None


BALL_VALVE = ValveTypeConfig(
    key="ball",
    # Nested under a "Ball Valve" subgroup (group Pune), so the menu reads
    # Pune › Ball Valve › Metal / Plastic instead of "Ball Valve" and
    # "Plastic Ball Valve" as two top-level rows. Same nesting pattern as the
    # Butterfly variants. Label is just the variant ("Metal"); the subgroup
    # supplies the "Ball Valve" prefix (result-panel heading shows both).
    # key stays "ball" so existing ?type=ball deep-links keep working.
    label="Metal",
    category="Valves",
    group="Pune",
    subgroup="Ball Valve",
    # file_substring is the FULL metal-master stem, NOT just "Ball Valve":
    # find_catalog_file() picks the most-recently-modified name match, and the
    # sibling "Ball Valve Plastic Data Sheet Structure" file is newer — a bare
    # "Ball Valve" substring would let the plastic file hijack this metal config
    # (its sheets carry no "BV NEW CODEING" marker → load crash). path_contains
    # pins it to Pune for good measure.
    file_substring="Ball Valve Data Sheet Structure NEW OG",
    path_contains="Pune",
    sheet_marker="BV NEW CODEING",
    # Recommended actuators per SKU. As of 2026-05-27, the richer source is
    # the dashboard file's "Ball Valve Actuator combination" sheet — it carries
    # pressure-specific labels (3.5/4/5.5 bar) for the 9 pneumatic positions
    # AND the 2 electric positions. The enrichment below pulls source cols
    # 49-59 into master cols 100-110, and these entries read the enriched cols.
    # Dedup is by (target_type, model), so if the same model is valid at
    # multiple pressures (common: ACT-050D works at 3.5, 4, AND 5.5 bar), the
    # user sees one chip with the first matching label.
    paired_actuators=(
        # Pneumatic Double Acting — 3 pressure variants
        PairedActuator(model_col=100, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting @ 3.5 bar"),
        PairedActuator(model_col=101, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting @ 4 bar"),
        PairedActuator(model_col=102, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting @ 5.5 bar"),
        # Pneumatic Spring Return Fail-Close — 3 pressure variants
        PairedActuator(model_col=103, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 3.5 bar"),
        PairedActuator(model_col=104, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 4 bar"),
        PairedActuator(model_col=105, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 5.5 bar"),
        # Pneumatic Spring Return Fail-Open — 3 pressure variants
        PairedActuator(model_col=106, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 3.5 bar"),
        PairedActuator(model_col=107, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 4 bar"),
        PairedActuator(model_col=108, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 5.5 bar"),
        # Electric — no sub-text on chips (user's choice which to pick).
        # Plain "Electric" labels are stripped to empty by the JS prefix
        # regex, so each chip shows just the model code with no sub-line.
        PairedActuator(model_col=109, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
        PairedActuator(model_col=110, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
        # Per-series files (cols 49-59 -> 111-121): the COMPLETE alternative
        # actuator set. The dashboard's fixed 3.5/4/5.5-bar slots above drop some
        # valid 3rd-alternative models (verified: 10 models across 140 SKUs, e.g.
        # ACT-063SR07). Dedup by (target_type, model) at resolve time means these
        # only ADD chips for models not already surfaced above — no duplicates.
        PairedActuator(model_col=111, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting"),
        PairedActuator(model_col=112, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting"),
        PairedActuator(model_col=113, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Double Acting"),
        PairedActuator(model_col=114, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close"),
        PairedActuator(model_col=115, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close"),
        PairedActuator(model_col=116, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close"),
        PairedActuator(model_col=117, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open"),
        PairedActuator(model_col=118, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open"),
        PairedActuator(model_col=119, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open"),
        PairedActuator(model_col=120, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
        PairedActuator(model_col=121, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
    ),
    cascade=[
        ("series",          4,  "Series"),
        ("size",            5,  "Valve Size"),
        ("body_material",   6,  "Body Material"),
        ("ball_material",   7,  "Ball Material"),
        ("seat_material",   8,  "Seat Material"),
        ("characteristics", 9,  "Characteristics"),
        ("end_connection", 10,  "End Connections"),
        ("ball_type",      22,  "Ball Type"),
    ],
    detail_columns=[
        (1, "Bare Valve Code"), (2, "Catalogue Code"), (3, "Make"),
        (11, "Valve Type"), (12, "Design Standard"), (13, "Face to Face"),
        (14, "Port Size"), (15, "No. of Ports"), (16, "Valve Kv (m³/hr)"),
        (17, "Body Style"), (18, "Flow Direction"), (19, "End Piece Material"),
        (20, "Type of Bonnet"), (21, "Stem Material"),
        (23, "Gland Packing"), (24, "Body Packing"),
        (25, "Flange Dimensions"), (26, "Flange Drilling"),
        (27, "Pressure Rating"), (28, "Operating Temp Range (°C)"),
        (29, "Hardware"), (30, "Valve Paint"),
        (31, "Testing Standard"), (32, "Leakage Class"),
        (33, "Body Test Pressure (barg)"), (34, "Body Test Media"),
        (35, "Seat Leakage Test Pressure (barg)"), (36, "Seat Leakage Test Media"),
        (37, "Product Group"), (38, "Certification"),
        (39, "BTO"), (40, "ETO"), (41, "BTC"), (42, "ETC"), (43, "Run"),
        (44, "Top PCD"), (45, "Stem Shape"), (46, "Stem Dimension"),
        (47, "Stem Orientation"), (48, "Stem Protrusion (mm)"),
        (49, "Double Acting Actuator 1"), (50, "Double Acting Actuator 2"),
        (51, "Double Acting Actuator 3"),
        (52, "Single Acting Fail-Safe Close 1"),
        (53, "Single Acting Fail-Safe Close 2"),
        (54, "Single Acting Fail-Safe Close 3"),
        (55, "Single Acting Fail-Safe Open 1"),
        (56, "Single Acting Fail-Safe Open 2"),
        (57, "Single Acting Fail-Safe Open 3"),
        # Enriched-in cols using descriptive labels from the dashboard file's
        # "Ball Valve Actuator combination" sheet (3.5 / 4 / 5.5 bar variants).
        (100, "Double Acting Actuator @ 3.5 bar"),
        (101, "Double Acting Actuator @ 4 bar"),
        (102, "Double Acting Actuator @ 5.5 bar"),
        (103, "Single Acting Fail-Safe Close @ 3.5 bar"),
        (104, "Single Acting Fail-Safe Close @ 4 bar"),
        (105, "Single Acting Fail-Safe Close @ 5.5 bar"),
        (106, "Single Acting Fail-Safe Open @ 3.5 bar"),
        (107, "Single Acting Fail-Safe Open @ 4 bar"),
        (108, "Single Acting Fail-Safe Open @ 5.5 bar"),
        (109, "Electric Actuator 1"), (110, "Electric Actuator 2"),
        # Per-series files (cols 49-59 -> 111-121): the complete alternative
        # actuator set, surfacing models the dashboard's pressure slots omit.
        (111, "Double Acting Actuator (alt 1)"), (112, "Double Acting Actuator (alt 2)"),
        (113, "Double Acting Actuator (alt 3)"),
        (114, "Single Acting Fail-Safe Close (alt 1)"),
        (115, "Single Acting Fail-Safe Close (alt 2)"),
        (116, "Single Acting Fail-Safe Close (alt 3)"),
        (117, "Single Acting Fail-Safe Open (alt 1)"),
        (118, "Single Acting Fail-Safe Open (alt 2)"),
        (119, "Single Acting Fail-Safe Open (alt 3)"),
        (120, "Electric Actuator (alt 1)"), (121, "Electric Actuator (alt 2)"),
    ],
    # Enrichment sources, merged in after the main load:
    #  1. The dashboard's "Ball Valve Actuator combination" sheet -> cols 100-110
    #     (9 pneumatic at 3.5/4/5.5 bar + 2 electric), pressure-labelled.
    #  2. The 4 per-series files -> NEW cols 111-121 (clear of 100-110). These
    #     carry the COMPLETE alternative actuator set; the union (deduped at
    #     resolve time) recovers models the dashboard's fixed pressure slots drop
    #     (verified: 10 models across 140 SKUs).
    enrichment_sources=(
        EnrichmentSource(
            file_substring="Valve_Code_Selector_Dashboard",
            sheet_marker="Ball Valve Actuator combination",
            source_key_col=1, master_key_col=1,
            columns=(
                (49, 100), (50, 101), (51, 102),
                (52, 103), (53, 104), (54, 105),
                (55, 106), (56, 107), (57, 108),
                (58, 109), (59, 110),
            ),
        ),
        EnrichmentSource(
            file_substring="2030F BV NEW CODEING", sheet_marker="2030F BV NEW CODEING",
            source_key_col=1, master_key_col=1,
            columns=((49,111),(50,112),(51,113),(52,114),(53,115),(54,116),(55,117),(56,118),(57,119),(58,120),(59,121)),
        ),
        EnrichmentSource(
            file_substring="2060F BV NEW CODEING", sheet_marker="2060F BV NEW CODEING",
            source_key_col=1, master_key_col=1,
            columns=((49,111),(50,112),(51,113),(52,114),(53,115),(54,116),(55,117),(56,118),(57,119),(58,120),(59,121)),
        ),
        EnrichmentSource(
            file_substring="2070F BV NEW CODEING", sheet_marker="2070F BV NEW CODEING",
            source_key_col=1, master_key_col=1,
            columns=((49,111),(50,112),(51,113),(52,114),(53,115),(54,116),(55,117),(56,118),(57,119),(58,120),(59,121)),
        ),
        EnrichmentSource(
            file_substring="2090F BV NEW CODEING", sheet_marker="2090F BV NEW CODEING",
            source_key_col=1, master_key_col=1,
            columns=((49,111),(50,112),(51,113),(52,114),(53,115),(54,116),(55,117),(56,118),(57,119),(58,120),(59,121)),
        ),
    ),
)


BALL_VALVE_PLASTIC = ValveTypeConfig(
    key="ball_plastic",
    # Sibling of BALL_VALVE under the "Ball Valve" subgroup (group Pune); label is
    # just the variant so the menu reads Pune › Ball Valve › Metal / Plastic.
    label="Plastic",
    category="Valves",
    group="Pune",
    subgroup="Ball Valve",
    # Series 2072K (U-PVC / C-PVC / PPH), 135 SKUs. Column layout is POSITIONALLY
    # IDENTICAL to the metal master, so the cascade/detail mirror BALL_VALVE.
    # sheet_marker "BV PLASTIC CODING" matches "2072K BV PLASTIC CODING" and any
    # future "<series> BV PLASTIC CODING" sheet. path_contains pins to Pune.
    file_substring="Ball Valve Plastic",
    path_contains="Pune",
    sheet_marker="BV PLASTIC CODING",
    # No FOS card: BTO (c39) is empty across all 135 rows, so torque/Factor-of-
    # Safety can't be computed — showing the card would leave it permanently blank.
    show_bto_fos=False,
    # Recommended actuators are baked into the sheet (unlike the metal master,
    # which enriches them from a separate dashboard). Plastic has only Spring-
    # Return Fail-Close (c49-51) + Fail-Open (c52-54) at 3.5/4/5.5 bar, all ACT-*
    # (pneumatic R&P), plus one Electric column (c55), all EA-* (electric rotary).
    # No Double-Acting columns exist. Dedup by (target_type, model) collapses a
    # model repeated across pressures to a single chip.
    paired_actuators=(
        PairedActuator(model_col=49, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 3.5 bar"),
        PairedActuator(model_col=50, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 4 bar"),
        PairedActuator(model_col=51, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 5.5 bar"),
        PairedActuator(model_col=52, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 3.5 bar"),
        PairedActuator(model_col=53, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 4 bar"),
        PairedActuator(model_col=54, target_type="pneumatic_rp", target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 5.5 bar"),
        PairedActuator(model_col=55, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
    ),
    cascade=[
        ("series",          4,  "Series"),
        ("size",            5,  "Valve Size"),
        ("body_material",   6,  "Body Material"),
        ("ball_material",   7,  "Ball Material"),
        ("seat_material",   8,  "Seat Material"),
        ("characteristics", 9,  "Characteristics"),
        ("end_connection", 10,  "End Connections"),
        ("ball_type",      22,  "Ball Type"),
    ],
    detail_columns=[
        (1, "Bare Valve Code"), (2, "Catalogue Code"), (3, "Make"),
        (11, "Valve Type"), (12, "Design Standard"), (13, "Face to Face"),
        (14, "Port Size"), (15, "No. of Ports"), (16, "Valve Kv (m³/hr)"),
        (17, "Body Style"), (18, "Flow Direction"), (19, "End Piece Material"),
        (20, "Type of Bonnet"), (21, "Stem Material"),
        (23, "O-Ring"), (24, "Body Packing"),
        (25, "Flange Dimensions"), (26, "Flange Drilling"),
        (27, "Pressure Rating"), (28, "Operating Temp Range (°C)"),
        (29, "Hardware"), (30, "Valve Paint"),
        (31, "Testing Standard"), (32, "Leakage Class"),
        (33, "Body Test Pressure (barg)"), (34, "Body Test Media"),
        (35, "Seat Leakage Test Pressure (barg)"), (36, "Seat Leakage Test Media"),
        (37, "Product Group"), (38, "Certification"),
        (44, "Top PCD"), (45, "Stem Shape"), (46, "Stem Dimension"),
        (47, "Stem Orientation"), (48, "Stem Protrusion (mm)"),
        # Baked-in recommended actuators (also drive the chips above).
        (49, "Spring Return Fail-Close @ 3.5 bar"),
        (50, "Spring Return Fail-Close @ 4 bar"),
        (51, "Spring Return Fail-Close @ 5.5 bar"),
        (52, "Spring Return Fail-Open @ 3.5 bar"),
        (53, "Spring Return Fail-Open @ 4 bar"),
        (54, "Spring Return Fail-Open @ 5.5 bar"),
        (55, "Electric Actuator"),
    ],
)


BUTTERFLY_VALVE = ValveTypeConfig(
    key="butterfly",
    # Nested under the "Butterfly Valve" subgroup (group Pune), so the menu reads
    # Pune › Butterfly Valve › Centric / Double Offset instead of two top-level
    # "Butterfly Valve (…)" rows. Label is just the variant; the subgroup supplies
    # the "Butterfly Valve" prefix (and the result-panel heading shows both).
    label="Centric",
    category="Valves",
    group="Pune",
    subgroup="Butterfly Valve",
    # R3 drop (2026-07-19): single pre-consolidated Sheet1 with 46,412 SKUs across
    # all 6 families (4020B/4022B/4023B + M bonnet variants). Actuator models are
    # baked directly into cols 49-59 (no per-series enrichment needed).
    # NOTE: Disc Type (c22) is blank for ~93 % of rows — kept out of the cascade
    # to avoid dead-ending most paths.
    file_substring="Butterfly Valve Centric R3",
    sheet_marker="Sheet1",
    paired_actuators=(
        # Cols 49-57: pneumatic actuators (DA×3, FC×3, FO×3) @ 3.5/4/5.5 bar.
        # Routing: ACT-* → Rack & Pinion, SYA-* → Scotch Yoke.
        PairedActuator(model_col=49, target_field="model",
                       label="Pneumatic — Double Acting @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=50, target_field="model",
                       label="Pneumatic — Double Acting @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=51, target_field="model",
                       label="Pneumatic — Double Acting @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=52, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=53, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=54, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=55, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=56, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=57, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        # Cols 58-59: electric actuator models (EA-* → electrical_rotary).
        # "(D)" = low torque / direct mount variant; "(E)" = larger variant.
        PairedActuator(model_col=58, target_type="electrical_rotary", target_field="model",
                       label="Electric (D)"),
        PairedActuator(model_col=59, target_type="electrical_rotary", target_field="model",
                       label="Electric (E)"),
    ),
    # Cascade follows AVCON Centric selection sequence (2026-07-19):
    # Series encodes body family + size (e.g. 4022B25 = Wafer, 1"). Valve Size
    # is omitted as a separate step because it is redundant with Series.
    cascade=[
        ("series",          4,  "Series"),
        ("body_material",   6,  "Body Material"),
        ("disc_material",   7,  "Disc Material"),
        ("seat_material",   8,  "Seat Material"),
        ("pressure_rating", 27, "Pressure Rating"),
        ("end_connection",  10, "End Connections"),
        ("face_to_face",    13, "Face to Face"),
        ("stem_material",   21, "Stem Material"),
        ("flange_drilling", 26, "Flange Drilling"),
    ],
    detail_columns=[
        (1, "Bare Valve Code"), (2, "Catalogue Code"), (3, "Make"),
        (5, "Valve Size"),
        (9, "Characteristics"),
        (11, "Valve Type"), (12, "Design Standard"),
        (14, "Port Size"), (15, "No. of Ports"), (16, "Valve Kv (m³/hr)"),
        (17, "Body Style"), (18, "Flow Direction"), (19, "Bonnet Material"),
        (20, "Type of Bonnet"),
        (23, "Gland Packing"), (24, "Body Packing"),
        (25, "Flange Dimensions"),
        (28, "Operating Temp Range (°C)"),
        (29, "Hardware"), (30, "Valve Paint"),
        (31, "Testing Standard"), (32, "Leakage Class"),
        (33, "Body Test Pressure (barg)"), (34, "Body Test Media"),
        (35, "Seat Leakage Test Pressure (barg)"), (36, "Seat Leakage Test Media"),
        (37, "Product Group"), (38, "Certification"),
        (39, "BTO"), (40, "ETO"), (41, "BTC"), (42, "ETC"), (43, "Run"),
        (44, "Top PCD"), (45, "Stem Shape"), (46, "Stem Dimension"),
        (47, "Stem Orientation"), (48, "Stem Protrusion (mm)"),
        # Actuator models baked in per-SKU (R3 layout: cols 49-59).
        (49, "Double Acting @ 3.5 bar"), (50, "Double Acting @ 4 bar"),
        (51, "Double Acting @ 5.5 bar"),
        (52, "Spring Return Fail-Close @ 3.5 bar"),
        (53, "Spring Return Fail-Close @ 4 bar"),
        (54, "Spring Return Fail-Close @ 5.5 bar"),
        (55, "Spring Return Fail-Open @ 3.5 bar"),
        (56, "Spring Return Fail-Open @ 4 bar"),
        (57, "Spring Return Fail-Open @ 5.5 bar"),
        (58, "Elect. Actuator Model (D)"), (59, "Elect. Actuator Model (E)"),
    ],
)


# Double-Offset (double-eccentric) butterfly — a SEPARATE family from the Centric
# type above (different disc geometry, its own 4026B/4027B/4028B series). Ships
# from a consolidated dashboard built by tools/consolidate_butterfly_double_offset.py:
# the drop's SKU sheet (Pharma-style +1 layout: leading Power-Query "Name" col, so
# Bare Valve Code is c2) plus size-based actuator SELECTION CHARTS. The tool joins
# each SKU's valve size to the charts and bakes 9 recommended actuator models into
# cols 60-68 (3 Double-Acting + 3 Spring-Return Fail-Close + 3 Fail-Open, at
# 3.5/4/5.5 bar). No electric chart exists for double-offset, so no electric chips.
# 18" SKUs (126 of 1,764) have no chart row -> no actuator recommendation (blank).
BUTTERFLY_DOUBLE_OFFSET = ValveTypeConfig(
    key="butterfly_double_offset",
    # Sibling of BUTTERFLY_VALVE under the "Butterfly Valve" subgroup (group Pune);
    # label is just the variant so the menu reads Pune › Butterfly Valve › Double Offset.
    label="Double Offset",
    category="Valves",
    group="Pune",
    subgroup="Butterfly Valve",
    file_substring="Butterfly Double Offset Dashboard",
    path_contains="Pune",
    sheet_marker="Butterfly Double Offset",
    # +1 vs the Centric layout: a leading Power-Query "Name" column (c1) pushes
    # Bare Valve Code to c2 and every other field down one (same shape as Pharma/
    # Control Valve). FOS torque uses BTO at c40 (ETO/BTC/ETC/Run are empty).
    primary_label="Bare Valve Code", primary_col=2,
    secondary_label="Catalogue Code", secondary_col=3,
    show_bto_fos=True, bto_col=40,
    # Disc Type (c23), Port Size (c15), No. of Ports (c16) are empty across all
    # 1,764 rows, so they're omitted from the cascade (a dropdown for them would
    # dead-end every selection) — same reasoning as the Centric "Disc Type" note.
    cascade=[
        ("series",          5,  "Series"),
        ("size",            6,  "Valve Size"),
        ("body_material",   7,  "Body Material"),
        ("disc_material",   8,  "Disc Material"),
        ("seat_material",   9,  "Seat Material"),
        # Stem Material (c22) added as a selector 2026-07-06 (user request). 100%
        # populated, 7 distinct grades; it's the field that otherwise left two
        # SKUs sharing all other picks (e.g. SS304 vs 17-4 PH), so surfacing it
        # lets the picker resolve to a single catalogue. Placed with the other
        # material dropdowns.
        ("stem_material",   22, "Stem Material"),
        ("characteristics", 10, "Characteristics"),
        ("end_connection",  11, "End Connections"),
    ],
    # Cols 60-68 hold the actuator models baked in by the consolidation tool;
    # they also feed the recommended-actuator chips (paired_actuators below).
    paired_actuators=(
        # Double Acting — 3 pressure variants (dedup by (target_type, model)
        # collapses a model repeated across pressures to one chip).
        PairedActuator(model_col=60, target_field="model",
                       label="Pneumatic — Double Acting @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=61, target_field="model",
                       label="Pneumatic — Double Acting @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=62, target_field="model",
                       label="Pneumatic — Double Acting @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        # Spring Return Fail-Close — 3 pressure variants
        PairedActuator(model_col=63, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=64, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=65, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        # Spring Return Fail-Open — 3 pressure variants
        PairedActuator(model_col=66, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 3.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=67, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 4 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=68, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open @ 5.5 bar",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
    ),
    detail_columns=[
        (2, "Bare Valve Code"), (3, "Catalogue Code"), (4, "Make"),
        (5, "Series"), (6, "Valve Size"), (7, "Body Material"),
        (8, "Disc Material"), (9, "Seat Material"), (10, "Characteristics"),
        (11, "End Connections"), (12, "Valve Type"), (13, "Design Standard"),
        (14, "Face to Face"), (15, "Port Size"), (16, "No. of Ports"),
        (17, "Valve Kv (m³/hr)"), (18, "Body Style"), (19, "Flow Direction"),
        (20, "Bonnet Material"), (21, "Type of Bonnet"), (22, "Stem Material"),
        (24, "Gland Packing"), (25, "Body Packing"),
        (26, "Flange Dimensions"), (27, "Flange Drilling"),
        (28, "Pressure Rating"), (29, "Operating Temp Range (°C)"),
        (30, "Hardware"), (31, "Valve Paint"),
        (32, "Testing Standard"), (33, "Leakage Class"),
        (34, "Body Test Pressure (barg)"), (35, "Body Test Media"),
        (36, "Seat Leakage Test Pressure (barg)"), (37, "Seat Leakage Test Media"),
        (38, "Product Group"), (39, "Certification"),
        (40, "BTO"), (41, "ETO"), (42, "BTC"), (43, "ETC"), (44, "Run"),
        (45, "Top PCD"), (46, "Stem Shape"), (47, "Stem Dimension"),
        (48, "Stem Orientation"), (49, "Stem Protrusion (mm)"),
        (50, "Additional Specification 1"), (51, "Additional Specification 2"),
        (52, "Additional Specification 3"), (53, "Additional Specification 4"),
        (54, "Additional Specification 5"), (55, "Bare Valve Weight (kg)"),
        # Actuator models baked in from the size-based selection charts.
        (60, "Double Acting @ 3.5 bar"), (61, "Double Acting @ 4 bar"),
        (62, "Double Acting @ 5.5 bar"),
        (63, "Spring Return Fail-Close @ 3.5 bar"),
        (64, "Spring Return Fail-Close @ 4 bar"),
        (65, "Spring Return Fail-Close @ 5.5 bar"),
        (66, "Spring Return Fail-Open @ 3.5 bar"),
        (67, "Spring Return Fail-Open @ 4 bar"),
        (68, "Spring Return Fail-Open @ 5.5 bar"),
    ],
)


PHARMA_VALVE = ValveTypeConfig(
    key="pharma",
    label="Pharma Valve",
    category="Valves",
    subgroup="Mumbai",
    # The catalog lives in the 'Pharma' sheet of a Valve_Code_Selector_Dashboard
    # .xlsx whose NAME collides with the Pune ball-valve dashboard, so we scope
    # the file search to the Mumbai subfolder.
    file_substring="Valve_Code_Selector_Dashboard",
    path_contains="Mumbai",
    sheet_marker="Pharma",
    # Column layout is shifted +1 vs the Ball Valve master: this sheet has a
    # leading Power-Query "Name" column (c1, "Table_…"), so Bare Valve Code is
    # c2, Series c5, etc. The 728 spacer rows with an empty Bare Valve Code are
    # dropped by the primary-col row filter in Catalog._load.
    primary_label="Bare Valve Code", primary_col=2,
    secondary_label="Catalogue Code", secondary_col=3,
    show_bto_fos=True, bto_col=40,
    # Recommended actuators are present IN this sheet (c50–60), so no enrichment
    # is needed (unlike ball/butterfly). Pneumatic models are ACT-* (Rack &
    # Pinion); prefix routing also handles any SYA-* (Scotch Yoke). Electric
    # models (c59–60) route to the rotary catalog. Dedup by (target_type, model)
    # at resolve time collapses repeats across the three positions.
    paired_actuators=(
        PairedActuator(model_col=50, target_field="model",
                       label="Pneumatic — Double Acting",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=51, target_field="model",
                       label="Pneumatic — Double Acting (alt 1)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=52, target_field="model",
                       label="Pneumatic — Double Acting (alt 2)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=53, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=54, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close (alt 1)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=55, target_field="model",
                       label="Pneumatic — Spring Return Fail-Close (alt 2)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=56, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=57, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open (alt 1)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        PairedActuator(model_col=58, target_field="model",
                       label="Pneumatic — Spring Return Fail-Open (alt 2)",
                       target_type_by_prefix=(("ACT", "pneumatic_rp"), ("SYA", "pneumatic_sy"))),
        # Linear actuator (HA-*, 5-6 bar) — a pneumatic operation type surfaced in
        # the SAME recommendation list as Double Acting / Spring Return. As of the
        # HA Series datasheet (2026-06-03) this IS a browsable family
        # (PNEUMATIC_HA below), so the chip is now clickable: target_type
        # "pneumatic_ha" routes the HA-* model into that catalog. Data present only
        # on the Pharma Dashboard SKUs that carry a linear model (col 122 via
        # extra_row_sources); all referenced models (HA-85/110/125 /NC) exist in
        # the HA catalog.
        PairedActuator(model_col=122, target_type="pneumatic_ha", target_field="model",
                       label="Pneumatic — Linear (5-6 bar)"),
        PairedActuator(model_col=59, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
        PairedActuator(model_col=60, target_type="electrical_rotary", target_field="model",
                       label="Electric"),
    ),
    # "Characteristics" (c10) is omitted from the cascade — it has a single value
    # ("ONF") across all rows, so a dropdown for it would be dead weight (same
    # reasoning the Electrical Rotary config uses). It still appears in details.
    cascade=[
        ("series",          5,  "Series"),
        ("size",            6,  "Valve Size"),
        ("body_material",   7,  "Body Material"),
        ("ball_material",   8,  "Ball Material"),
        ("seat_material",   9,  "Seat Material"),
        ("end_connection", 11,  "End Connections"),
        ("ball_type",      23,  "Ball Type"),
    ],
    detail_columns=[
        (2, "Bare Valve Code"), (3, "Catalogue Code"), (4, "Make"),
        (5, "Series"), (6, "Valve Size"),
        (7, "Body Material"), (8, "Ball Material"), (9, "Seat Material"),
        (10, "Characteristics"), (11, "End Connections"),
        (12, "Valve Type"), (13, "Design Standard"), (14, "Face to Face"),
        (15, "Port Size"), (16, "No. of Ports"), (17, "Valve Kv (m³/hr)"),
        (18, "Body Style"), (19, "Flow Direction"), (20, "End Piece Material"),
        (21, "Type of Bonnet"), (22, "Stem Material"), (23, "Ball Type"),
        (24, "Gland Packing"), (25, "Body Packing"),
        (26, "Flange Dimensions"), (27, "Flange Drilling"),
        (28, "Pressure Rating"), (29, "Operating Temp Range (°C)"),
        (30, "Hardware"), (31, "Valve Paint"),
        (32, "Testing Standard"), (33, "Leakage Class"),
        (34, "Body Test Pressure (barg)"), (35, "Body Test Media"),
        (36, "Seat Leakage Test Pressure (barg)"), (37, "Seat Leakage Test Media"),
        (38, "Product Group"), (39, "Certification"),
        (40, "BTO"), (41, "ETO"), (42, "BTC"), (43, "ETC"), (44, "Run"),
        (45, "Top PCD"), (46, "Stem Shape"), (47, "Stem Dimension"),
        (48, "Stem Orientation"), (49, "Stem Protrusion (mm)"),
        (50, "Double Acting Actuator 1"), (51, "Double Acting Actuator 2"),
        (52, "Double Acting Actuator 3"),
        (53, "Single Acting Fail-Safe Close 1"),
        (54, "Single Acting Fail-Safe Close 2"),
        (55, "Single Acting Fail-Safe Close 3"),
        (56, "Single Acting Fail-Safe Open 1"),
        (57, "Single Acting Fail-Safe Open 2"),
        (58, "Single Acting Fail-Safe Open 3"),
        (59, "Electric Actuator 1"), (60, "Electric Actuator 2"),
        # From the Pharma Dashboard extra-row source (no equivalent in the main
        # 'Pharma' sheet). Linear (122) ALSO surfaces as a 'Linear' chip in the
        # Recommended Actuator list (paired_actuators above) — it's a pneumatic
        # operation type, not just a spec; it stays here too so the full
        # attribute dump is complete. Manual (123) is detail-only.
        (122, "Linear Actuator (5-6 bar)"), (123, "Manual Operator"),
    ],
    extra_row_sources=(
        # Fold in the 162 Pharma SKUs that exist ONLY in Pharma Dashboard.xlsx
        # ('Ball Valve Data Sheet Structure' sheet, ball layout: c1=Bare Code, so
        # base cols map +1 into the 'Pharma' layout). Pneumatic/electric actuators
        # land in cols 50-60 (surfaced as chips by paired_actuators above);
        # Linear/Manual land in 122/123 (detail only). key_pattern skips the
        # sheet's embedded header rows.
        ExtraRowSource(
            file_substring="Pharma Dashboard",
            sheet_marker="Ball Valve Data Sheet Structure",
            path_contains="Mumbai",
            key_col=1, master_key_col=2,
            key_pattern=r"^[0-9]{3,4}[A-Z]",
            column_map=(
                (1,2),(2,3),(3,4),(4,5),(5,6),(6,7),(7,8),(8,9),(9,10),(10,11),
                (11,12),(12,13),(13,14),(14,15),(15,16),(16,17),(17,18),(18,19),
                (19,20),(20,21),(21,22),(22,23),(23,24),(24,25),(25,26),(26,27),
                (27,28),(28,29),(29,30),(30,31),(31,32),(32,33),(33,34),(34,35),
                (35,36),(36,37),(37,38),(38,39),(39,40),(40,41),(41,42),(42,43),
                (43,44),(44,45),(45,46),(46,47),(47,48),(48,49),
                (49,50),(50,51),(51,52),(52,53),(53,54),(54,55),(55,56),(56,57),(57,58),
                (59,59),(60,60),
                (58,122),(61,123),
            ),
        ),
    ),
)


# ---- Actuators --------------------------------------------------------------
# All three actuator catalogs share a similar layout:
#   col 1 = CODE          (the SKU)
#   col 5 = Model         (human-friendly identifier)
#   col 4 = Type          (DA / Spring Return / Rotary motorised / …)
#   per-type specs from col 6 onwards
# Headline cards are Code + Model. BTO/FOS doesn't apply.

PNEUMATIC_RACK_PINION = ValveTypeConfig(
    key="pneumatic_rp",
    label="Rack & Pinion",
    category="Actuators",
    subgroup="Rotary",
    group="Pneumatic",
    file_substring="Rack & Pinion",
    sheet_marker="Rack & Pinion ACT",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # The 4 fields that uniquely determine a Code per AVCON's workflow.
    # Verified 2026-05-26: {Model + Body + Shaft Female + O-Ring} disambiguates
    # 1,618 of 1,619 combinations (99.9%); only ACT-075SR12+Aluminium+SQ.17+
    # Viton Low resolves to two codes (RPAS0716 / RPAS0724) which differ on
    # Temperature Rating only. Other attributes (Type, Springs, Pneumatic
    # Connections, Application, Certification) stay in detail_columns so they
    # still display in the resolved-actuator panel.
    cascade=[
        ("model",            5, "Model"),
        ("body_material",    8, "Body Material"),
        ("shaft_female",    21, "Shaft Female"),
        ("oring_material",   9, "O-Ring Material"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (6, "No. of Springs"), (7, "Movement"),
        (8, "Body Material"), (9, "O-Ring Material"), (10, "Temperature Rating"),
        (11, "Min Pneumatic Supply"), (12, "Max Pneumatic Supply"),
        (13, "Pneumatic Connections"), (14, "Actuator Orientation"),
        (15, "Application"), (16, "Painting"), (17, "Certification"),
        (18, "Catalogue Code"), (21, "Shaft Female"), (22, "PCD"),
        (23, "Torque @ 2.5 bar (Break)"), (24, "Torque @ 2.5 bar (End)"),
        (25, "Torque @ 3 bar (Break)"),   (26, "Torque @ 3 bar (End)"),
        (27, "Torque @ 3.5 bar (Break)"), (28, "Torque @ 3.5 bar (End)"),
        (29, "Torque @ 4 bar (Break)"),   (30, "Torque @ 4 bar (End)"),
        (31, "Torque @ 4.5 bar (Break)"), (32, "Torque @ 4.5 bar (End)"),
        (33, "Torque @ 5 bar (Break)"),   (34, "Torque @ 5 bar (End)"),
        (35, "Torque @ 5.5 bar (Break)"), (36, "Torque @ 5.5 bar (End)"),
        (37, "Torque @ 6 bar (Break)"),   (38, "Torque @ 6 bar (End)"),
        (39, "Torque @ 7 bar (Break)"),   (40, "Torque @ 7 bar (End)"),
        (41, "Torque @ 8 bar (Break)"),   (42, "Torque @ 8 bar (End)"),
        (43, "Spring Break Torque"), (44, "Spring End Torque"),
    ],
)


PNEUMATIC_SCOTCH_YOKE = ValveTypeConfig(
    key="pneumatic_sy",
    label="Scotch Yoke",
    category="Actuators",
    subgroup="Rotary",
    group="Pneumatic",
    file_substring="Scotch Yoke",
    sheet_marker="Scotch Yoke Actuator SYA",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # Same 4-field shape as R&P. Verified 2026-05-26: 479 of 480 combos
    # (99.8%) resolve to one Code; one duplicate (SYAS0604 / SYAS0606) shows
    # no difference in any declared column — likely a data-quality issue
    # worth flagging engineering.
    cascade=[
        ("model",            5, "Model"),
        ("body_material",    8, "Body Material"),
        ("shaft_female",    21, "Shaft Female"),
        ("oring_material",   9, "O-Ring Material"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (6, "Spring Model"), (7, "Movement"),
        (8, "Body Material"), (9, "O-Ring Material"), (10, "Temperature Rating"),
        (11, "Min Pneumatic Supply"), (12, "Max Pneumatic Supply"),
        (13, "Pneumatic Connections"), (14, "Actuator Orientation"),
        (15, "Application"), (16, "Painting"), (17, "Certification"),
        (21, "Shaft Female"), (22, "PCD"),
        (23, "Torque @ 3 bar (Break)"),   (25, "Torque @ 3 bar (End)"),
        (26, "Torque @ 3.5 bar (Break)"), (28, "Torque @ 3.5 bar (End)"),
        (29, "Torque @ 4 bar (Break)"),   (31, "Torque @ 4 bar (End)"),
        (32, "Torque @ 4.5 bar (Break)"), (34, "Torque @ 4.5 bar (End)"),
        (35, "Torque @ 5 bar (Break)"),   (37, "Torque @ 5 bar (End)"),
        (38, "Torque @ 5.5 bar (Break)"), (40, "Torque @ 5.5 bar (End)"),
        (41, "Torque @ 6 bar (Break)"),   (43, "Torque @ 6 bar (End)"),
        (44, "Torque @ 7 bar (Break)"),   (46, "Torque @ 7 bar (End)"),
        (47, "Torque @ 8 bar (Break)"),   (49, "Torque @ 8 bar (End)"),
        (50, "Spring Break Torque"), (52, "Spring End Torque"),
    ],
)


ELECTRICAL_ROTARY = ValveTypeConfig(
    key="electrical_rotary",
    label="Rotary",
    category="Actuators",
    subgroup="Electrical",
    # Was "Electrical Actuator" — too broad: it also matched the newer
    # "Electrical Actuator Linear..." file, which (being newest) hijacked this
    # catalog. Anchored to "Rotory" (the source-file spelling) so Rotary and
    # Linear load as separate families. See ELECTRICAL_LINEAR below.
    file_substring="Electrical Actuator Rotory",
    sheet_marker="Electrical Actuator",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # All other catalog fields (Type, Body Material, Enclosure, Application,
    # Certification) have exactly 1 distinct value across all 96 rows, so
    # they're omitted from the cascade — they appear in the detail table only.
    cascade=[
        ("model",   5, "Model"),
        ("voltage", 6, "Voltage"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (6, "Voltage"), (7, "Movement"),
        (8, "Body Material"), (9, "O-Ring Material"), (10, "Temperature Rating"),
        (11, "Power Consumption (W)"), (12, "Open-to-Close Time (s)"),
        (13, "Enclosure Protection"), (14, "Actuator Orientation"),
        (15, "Application"), (16, "Painting"), (17, "Certification"),
        (21, "Shaft Female"), (22, "PCD"),
        (23, "Torque (N·m)"), (24, "Weight (kg)"),
    ],
    # NOTE: Actuator.xlsx's 'Working' sheet contains 2 codes not in this catalog
    # (E021GN02, E021GX02 = EA-21/E "with Potentiometer" @ 110 VAC 50/60 Hz). They
    # were NOT folded in: E021GN02 collides with E021GN01 on Model+Voltage (the
    # only cascade fields), and the source is a combined working template whose
    # columns don't map cleanly to this catalog. If these variants are needed,
    # add them to the authoritative datasheet with a distinguishing attribute.
)


ELECTRICAL_LINEAR = ValveTypeConfig(
    key="electrical_linear",
    label="Linear",
    category="Actuators",
    subgroup="Electrical",
    file_substring="Electrical Actuator Linear",
    sheet_marker="Electrical Actuator",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 40 rows. Model (8 distinct) + Voltage (5) are the only meaningful
    # disambiguators; "Additional Sp 1" (ONF vs with-Potentiometer) added so
    # potentiometer variants don't collide on Model+Voltage.
    cascade=[
        ("model",   5, "Model"),
        ("voltage", 6, "Voltage"),
        ("variant", 18, "Variant"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (6, "Voltage"), (8, "Body Material"), (10, "Temperature Rating"),
        (11, "Power Consumption (W)"), (12, "Open-to-Close Time (s)"),
        (13, "Enclosure Protection"), (15, "Application"), (16, "Painting"),
        (17, "Certification"), (18, "Variant"),
        (22, "Maximum Stroke (mm)"), (24, "Weight (kg)"), (25, "Force (kN)"),
    ],
)


PNEUMATIC_CY = ValveTypeConfig(
    key="pneumatic_cy",
    label="CY Series",
    category="Actuators",
    subgroup="Cylinder",
    group="Pneumatic",
    file_substring="CY Series",
    sheet_marker="Sheet2",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 98 rows; Model is effectively unique. Action (NC/NO) is the only other
    # varying field — everything else is uniform across the series.
    cascade=[
        ("model",  5, "Model"),
        ("action", 4, "Action"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator Type"), (3, "Make"), (4, "Action"),
        (5, "Model"), (7, "Movement"), (8, "Body/Bonnet Material"),
        (9, "O-Ring Material"), (10, "Temperature Rating"),
        (11, "Min Pneumatic Supply"), (12, "Max Pneumatic Supply"),
        (13, "Pneumatic Connections"), (18, "Leakage Class"),
    ],
)


PNEUMATIC_K = ValveTypeConfig(
    key="pneumatic_k",
    label="K Series",
    category="Actuators",
    subgroup="Cylinder",
    group="Pneumatic",
    file_substring="K Series",
    sheet_marker="K-Series",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 143 rows; Model is unique. Action (NC/NO) and Body Material (3 grades)
    # narrow the model list. Col 2 carries 4 handwheel-variant descriptions.
    cascade=[
        ("model",         5, "Model"),
        ("action",        4, "Action"),
        ("body_material", 8, "Body Material"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Action"), (5, "Model"),
        (7, "Movement"), (8, "Body/Bonnet Material"), (9, "O-Ring Material"),
        (10, "Temperature Rating"), (11, "Min Pneumatic Supply"),
        (12, "Max Pneumatic Supply"), (13, "Pneumatic Connections"),
    ],
)


PNEUMATIC_M = ValveTypeConfig(
    key="pneumatic_m",
    label="M Series",
    category="Actuators",
    subgroup="Cylinder",
    group="Pneumatic",
    file_substring="M Series",
    sheet_marker="Sheet2",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 8 rows; only Model varies (all other columns uniform).
    cascade=[
        ("model", 5, "Model"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (7, "Movement"), (8, "Body/Bonnet Material"), (9, "O-Ring Material"),
        (10, "Temperature Rating"), (11, "Min Pneumatic Supply"),
        (12, "Max Pneumatic Supply"), (13, "Pneumatic Connections"),
    ],
)


MANUAL_HANDWHEEL = ValveTypeConfig(
    key="manual_handwheel",
    label="Manual Handwheel (MPW)",
    category="Actuators",
    subgroup="Manual",
    file_substring="Manual Handwheel (MPW)",  # specific: don't match "MHL Manual Handwheel Lever"
    sheet_marker="Sheet2",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 44 rows; Model is unique. Type (Plastic vs Stainless Steel handwheel) is
    # the one other varying field. This family realizes the planned "Manual"
    # subgroup placeholder under Actuators.
    cascade=[
        ("model",   5, "Model"),
        ("variant", 2, "Type"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator"), (3, "Make"), (4, "Type"), (5, "Model"),
        (7, "Movement"), (8, "Body/Bonnet Material"), (9, "O-Ring Material"),
        (10, "Temperature Rating"),
    ],
)


PNEUMATIC_HA = ValveTypeConfig(
    key="pneumatic_ha",
    label="Linear (HA Series)",
    category="Actuators",
    subgroup="Linear",
    group="Pneumatic",
    file_substring="HA Series",
    sheet_marker="Sheet2",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=5,
    show_bto_fos=False,
    # 6 rows (HA-85/110/125 in NC & NO). Same CY/M-Series layout. Model is the
    # unique key; Action (NC/NO) is the only other varying field. This is the
    # "Linear (5-6 bar)" actuator referenced by PHARMA_VALVE's recommendation
    # chip — wiring that chip to target_type="pneumatic_ha" makes it clickable.
    cascade=[
        ("model",  5, "Model"),
        ("action", 4, "Action"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Actuator Type"), (3, "Make"), (4, "Action"),
        (5, "Model"), (7, "Movement"), (8, "Body/Bonnet Material"),
        (9, "O-Ring Material"), (10, "Temperature Rating"),
        (11, "Min Pneumatic Supply"), (12, "Max Pneumatic Supply"),
        (13, "Pneumatic Connections"), (18, "Leakage Class"),
    ],
)


CONTROL_VALVE = ValveTypeConfig(
    key="control_valve",
    label="Control Valve",
    category="Valves",
    group="Pune",
    file_substring="Control Valve Dashboard",
    path_contains="Pune",
    sheet_marker="Control Valve",
    # Power-Query layout: c1 is the table-name column ("…CV"), so Bare Valve
    # Code is c2 and Catalogue is c3 (same shape as PHARMA_VALVE).
    primary_label="Bare Valve Code", primary_col=2,
    secondary_label="Catalogue Code", secondary_col=3,
    show_bto_fos=False,
    # 17,681 rows → 1,831 Catalogue specs. This cascade resolves the Catalogue
    # code UNIQUELY (0 ambiguous). It was trimmed 16→8 (2026-06-11), then 3 fields
    # were re-added by user request (8→11, 2026-06-12; see UPDATE note below).
    # The 2026-06-11 trim rationale (still valid — adding selectors only refines):
    # the dropped 8 fields are each either 1:1 with a kept field or fully implied
    # by Series, so removing them as SELECTORS doesn't reintroduce ambiguity —
    #   valve_type/port_size/num_ports/body_style/plug_type  <- implied by series
    #   bonnet_material                                       <- 1:1 with body_material
    #   stem_material                                         <- implied by trim_material
    #   flow_direction                                        <- redundant for uniqueness
    # All 8 dropped fields STILL appear in detail_columns (shown in the result),
    # so no spec information is lost — only the number of dropdowns shrinks.
    # `Type Of Bonnet` is load-bearing and CANNOT be dropped: it's the only field
    # that distinguishes otherwise-identical catalogues (e.g. Bellow Seal vs
    # Standard) — verified that no 0-ambiguous set exists without it.
    # NOTE: cost/pricing columns (56-59) are deliberately excluded from cascade
    # AND detail so they never reach the public JSON.
    # UPDATE 2026-06-12 (user request): re-added 3 dropdowns after End Connections
    # — Face to Face (14), Port Size (15), Certification (39) — and moved the
    # existing Type Of Bonnet (21) ahead of Certification. Now 11 fields. Adding
    # selectors can only refine matches, so 0-ambiguous still holds (re-verified).
    # Caveat: Certification (39) is ~37% populated; its dropdown is empty for
    # valves with no cert data, but resolution is progressive so it never
    # dead-ends the picker (verified: app.js resolves on current picks).
    # UPDATE 2026-06-14 (user request): re-added Flow Direction (19) after
    # Characteristics — it was dropped in the 16→8 cut as "redundant for
    # uniqueness", but the user wants it selectable. 100% populated; for 2-way
    # series it's uniformly "Flow to Open" (single-option dropdown), only the
    # 3-way series (5016/5066) offer Diverting/Mixing. Now 12 fields; monotonic,
    # 0-ambiguous re-verified.
    cascade=[
        ("series",          5, "Series"),
        ("size",            6, "Valve Size"),
        ("body_material",   7, "Body Material"),
        ("trim_material",   8, "Trim Material"),
        ("seat_material",   9, "Seat Material"),
        ("characteristics", 10, "Characteristics"),
        ("flow_direction",  19, "Flow Direction"),
        ("end_connection",  11, "End Connections"),
        ("face_to_face",    14, "Face to Face"),
        ("port_size",       15, "Port Size (mm)"),
        ("valve_kv",        17, "Valve Kv (m³/hr)"),
        ("bonnet_type",     21, "Type Of Bonnet"),
        ("certification",   39, "Certification"),
    ],
    # Per-series cascade field set + ORDER (user request 2026-07-12). The
    # override list is the exact ordered list of visible fields; app.js reorders
    # the dropdowns to match. All sets verified 0-ambiguous per series. 5061A/
    # 5066A are pinned to the historical 12 fields so adding valve_kv to the base
    # cascade (for 5012) does not leak an extra dropdown into them.
    cascade_override_key="series",
    cascade_overrides={
        # 5016A/5016B: 13 fields. valve_kv 100% populated; certification 0% (shows
        # "Not Applicable" only — included for consistency with other series).
        "5016A": ["series", "size", "body_material", "trim_material", "seat_material",
                  "characteristics", "flow_direction", "end_connection", "face_to_face",
                  "port_size", "valve_kv", "bonnet_type", "certification"],
        "5016B": ["series", "size", "body_material", "trim_material", "seat_material",
                  "characteristics", "flow_direction", "end_connection", "face_to_face",
                  "port_size", "valve_kv", "bonnet_type", "certification"],
        # 5012A/5012B: 12 fields — drops Valve Size, adds Valve Kv; Flow
        # Direction moved after Port Size / Valve Kv (this exact order).
        "5012A": ["series", "body_material", "trim_material", "seat_material",
                  "characteristics", "end_connection", "face_to_face",
                  "port_size", "valve_kv", "flow_direction", "bonnet_type",
                  "certification"],
        "5012B": ["series", "body_material", "trim_material", "seat_material",
                  "characteristics", "end_connection", "face_to_face",
                  "port_size", "valve_kv", "flow_direction", "bonnet_type",
                  "certification"],
        # 5061A/5066A: unchanged historical 12 (base order, no Valve Kv).
        "5061A": ["series", "size", "body_material", "trim_material",
                  "seat_material", "characteristics", "flow_direction",
                  "end_connection", "face_to_face", "port_size", "bonnet_type",
                  "certification"],
        "5066A": ["series", "size", "body_material", "trim_material",
                  "seat_material", "characteristics", "flow_direction",
                  "end_connection", "face_to_face", "port_size", "bonnet_type",
                  "certification"],
    },
    detail_columns=[
        (2, "Bare Valve Code"), (3, "Catalogue Code"), (4, "Make"),
        (5, "Series"), (6, "Valve Size"), (7, "Body Material"),
        (8, "Trim Material"), (9, "Seat Material"), (10, "Characteristics"),
        (11, "End Connections"), (12, "Valve Type"), (13, "Design Standard"),
        (14, "Face to Face"), (15, "Port Size (mm)"), (16, "No. Of Ports"),
        (17, "Valve Kv (m³/hr)"), (18, "Body Style"), (19, "Flow Direction"),
        (20, "Bonnet Material"), (21, "Type Of Bonnet"), (22, "Stem Material"),
        (23, "Plug Type"), (24, "Gland Packing"), (25, "Body Packing"),
        (26, "Flange Dimensions"), (27, "Flange Drilling"), (28, "Pressure Rating"),
        (29, "Operating Temp Range (°C)"), (30, "Hardware"), (31, "Valve Paint"),
        (32, "Testing Standard"), (33, "Leakage Class"),
        (34, "Body Test Pressure (barg)"), (35, "Body Test Media"),
        (36, "Seat Leakage Test Pressure (barg)"), (37, "Seat Leakage Test Media"),
        (38, "Product Group"), (39, "Certification"), (40, "Thrust (kN)"),
        (41, "Torque (N·m)"), (42, "Mounting PCD"), (43, "Stroke (mm)"),
        (44, "Stem Diameter"),
        # cols 45-49: AVCON's R2 drop named these spec slots (were the generic
        # "Additional Specification 1-5", dropped in R1). Consolidation normalises
        # the series-varying source headers by position; see
        # tools/consolidate_control_valve.py SPEC_SLOTS.
        # col 49 (2026-06-14 re-drop): 5012A/B split control pressure by fail-safe
        # mode — col 48 holds the Fail-Safe-Close value, col 49 the Open value.
        # Populated only on 5012A/B; blank on the other series.
        (45, "Max Shut-off Pressure"),
        (46, "Fail Safe Close / A-Port Close"),
        (47, "Fail Safe Open / B-Port Close"),
        (48, "Control Pressure"),
        (49, "Control Pressure (Fail Safe Open)"),
        (55, "Bare Valve Weight (kg)"),
    ],
    # Recommended actuator (from AVCON's R2 drop, 2026-06-11). The Fail-Safe
    # columns name the MSD linear actuator the valve pairs with — but as a
    # model+variant ("MSD-200 E") and only on ~64% of rows (the rest hold a
    # pressure or 0). require_prefix="MSD-" drops the non-model cells; the
    # _MSD_FAMILY_PATTERN normaliser maps "MSD-200 E" -> the MSD `model`
    # family "MSD-200" for the picker prefill (bore/spring variant is lost —
    # user selects it in the MSD catalog). Cols 46/47 are also detail columns,
    # so they're already loaded. Close = normally-closed, Open = normally-open;
    # 2-way valves name the same model in both (dedup -> one card), 3-way valves
    # may differ (two cards).
    paired_actuators=(
        PairedActuator(
            model_col=46, target_field="model", target_type="pneumatic_msd",
            require_prefix="MSD-", label="Fail-Safe Close (Normally Closed)",
            series_labels={"5016A": "A Port Close", "5016B": "A Port Close"},
        ),
        PairedActuator(
            model_col=47, target_field="model", target_type="pneumatic_msd",
            require_prefix="MSD-", label="Fail-Safe Open (Normally Open)",
            series_labels={"5016A": "B Port Close", "5016B": "B Port Close"},
        ),
    ),
)


# MSD multi-spring diaphragm — a second PNEUMATIC LINEAR actuator (joins HA under
# Pneumatic › Linear). 250 rows in the "MSD ACT" sheet. A code is identified by
# Model + Type (NC/NO) + No. of Springs + Yoke Bore, matching the Catlouge form
# MSD-<model>-<spring letter>-<NC/NO>-<bore>. (Also the actuator the THW
# accessory pairs with — Suitable Actuator Model = MSD-*.)
PNEUMATIC_MSD = ValveTypeConfig(
    key="pneumatic_msd",
    label="MSD Series",
    category="Actuators",
    subgroup="Linear",
    group="Pneumatic",
    file_substring="MSD Actuator",
    sheet_marker="MSD ACT",
    primary_label="Code", primary_col=1,
    secondary_label="Catalogue Code", secondary_col=2,
    show_bto_fos=False,
    cascade=[
        ("model",      6, "Model"),
        ("action",     5, "Type"),
        ("springs",    7, "No. Of Springs"),
        ("yoke_bore", 23, "Yoke Bore Dia. (mm)"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Catalogue Code"), (3, "Actuator"), (4, "Make"),
        (5, "Type"), (6, "Model"), (7, "No. Of Springs"), (8, "Movement"),
        (9, "Top/Bottom Cover Material"), (10, "Diaphragm Material"),
        (11, "Temperature Rating"), (12, "Min Pneumatic Supply"),
        (13, "Max Pneumatic Supply"), (15, "Pneumatic Connections"),
        (16, "Yoke Material"), (23, "Yoke Bore Dia. (mm)"),
        (24, "MSD Actuator Force"), (25, "Stem Dia."), (27, "Stroke (mm)"),
    ],
)

# MHG manual gear box — a MANUAL actuator (alongside Manual Handwheel MPW). 41
# rows in the "MHG" sheet; Model is unique. Narrow by Make → Torque → Model.
MANUAL_GEARBOX = ValveTypeConfig(
    key="manual_gearbox",
    label="Gear Box (MHG)",
    category="Actuators",
    subgroup="Manual",
    file_substring="MHG Manual Gear Box",
    sheet_marker="MHG",
    primary_label="Code", primary_col=1,
    secondary_label="Model", secondary_col=3,
    show_bto_fos=False,
    cascade=[
        ("make",   2, "Make"),
        ("torque", 7, "Torque (Nm)"),
        ("model",  3, "Model"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Make"), (3, "Model"), (4, "Material of Construction"),
        (5, "Valve Side PCD"), (6, "Painting"), (7, "Torque (Nm)"),
        (8, "Handwheel Material"),
    ],
    # Shakti gear boxes (18 rows, MGS*/SE-*) dropped 2026-07-29 into
    # data/Accessories/. Schema-identical to the master (same 14 headers), so
    # columns map 1:1. path_contains is REQUIRED: the bare file_substring also
    # matches this catalog's OWN file ("MHG Manual Gear Box Data for New
    # Structure_R1.xlsx"), and picking that would add zero rows silently.
    extra_row_sources=(
        ExtraRowSource(
            file_substring="Gear Box Data for New Structure",
            sheet_marker="MHG",
            path_contains="Accessories",
            key_col=1, master_key_col=1,
            column_map=((1,1),(2,2),(3,3),(4,4),(5,5),(6,6),(7,7),(8,8),(9,9)),
        ),
    ),
)

# MHL manual handwheel lever — a MANUAL actuator. Only 2 rows ("Sheet1"); the
# sole varying field is Material of Construction (Carbon Steel vs SS316).
MANUAL_LEVER = ValveTypeConfig(
    key="manual_lever",
    label="Handwheel Lever (MHL)",
    category="Actuators",
    subgroup="Manual",
    file_substring="MHL Manual Handwheel Lever",
    sheet_marker="Sheet1",
    primary_label="Code", primary_col=1,
    secondary_label="Material", secondary_col=2,
    show_bto_fos=False,
    cascade=[
        ("material", 2, "Material of Construction"),
    ],
    detail_columns=[
        (1, "Code"), (2, "Material of Construction"),
    ],
)


VALVE_TYPES: list[ValveTypeConfig] = [
    BALL_VALVE,
    BALL_VALVE_PLASTIC,
    BUTTERFLY_VALVE,
    BUTTERFLY_DOUBLE_OFFSET,
    PHARMA_VALVE,
    CONTROL_VALVE,
    PNEUMATIC_RACK_PINION,
    PNEUMATIC_SCOTCH_YOKE,
    PNEUMATIC_CY,
    PNEUMATIC_K,
    PNEUMATIC_M,
    PNEUMATIC_HA,
    PNEUMATIC_MSD,
    ELECTRICAL_ROTARY,
    ELECTRICAL_LINEAR,
    MANUAL_HANDWHEEL,
    MANUAL_GEARBOX,
    MANUAL_LEVER,
]


# Subgroups that should appear in a category's picker menu even before any
# catalog populates them. Each entry: category -> ordered list of
# (subgroup_name, placeholder_text). Auto-superseded once a real family loads
# into the subgroup (see build_category_blocks). Only injected into categories
# that already have at least one loaded family.
PLANNED_SUBGROUPS: dict[str, list[tuple[str, str]]] = {
    "Valves": [("Mumbai", "Data pending — Mumbai catalog coming soon")],
    # Manual (SHL) is a distinct, hand-operated actuator family referenced by
    # Pharma valves; its catalog is pending (data/Actuator/Manual Actuator) and
    # shows as a placeholder subgroup until supplied.
    # Linear (HA-*, 5-6 bar) is NOT a browsable family — it's a pneumatic
    # operation type, surfaced as a chip in each valve's Recommended Actuator
    # list (see PHARMA_VALVE.paired_actuators), so it has no picker subgroup.
    "Actuators": [
        ("Manual", "Data pending — Manual actuator catalog coming soon"),
    ],
}


def _norm(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "")
        return s if s else None
    return v


import re

# TRANSITIONAL: source-data format drift between per-series files and the
# destination actuator catalogs. Two patterns observed (2026-05-26 +
# 2026-05-27 validation reports):
#   (1) Missing dash after letter prefix:
#         `SYA065300DAS` (per-series)  vs  `SYA-065300DAS` (SY catalog)
#   (2) Missing slash before trailing single letter:
#         `EA-21D` (per-series)  vs  `EA-21/D` (electrical_rotary catalog)
#         `QM-150D` (per-series)  vs  `QM-150/D` (electrical_rotary catalog)
# We normalize at load time so lookups succeed even when engineering hasn't
# yet cleaned the source files. Remove when source data is consistent.
_PAIRED_MODEL_DASH_PREFIXES = ("SYA",)
_EA_QM_SLASH_PATTERN = re.compile(r"^(EA|QM)-(\d+)([A-Z])$")
# Control valve "Fail Safe" cells name an MSD model + bore/spring variant letter
# ("MSD-200 E", "MSD-250D"); the pneumatic_msd catalog's `model` field is just
# the family ("MSD-200"). Strip the optional space + single trailing letter.
_MSD_FAMILY_PATTERN = re.compile(r"^(MSD-\d+)\s*[A-Za-z]?$")


def _normalize_paired_model(model: str) -> str:
    """Apply known format-drift corrections so the recommended-model string
    matches the canonical form used in the destination actuator catalog."""
    # Pattern 1: insert missing dash after prefix (SYA065 -> SYA-065)
    for prefix in _PAIRED_MODEL_DASH_PREFIXES:
        if model.startswith(prefix) and not model.startswith(f"{prefix}-"):
            model = f"{prefix}-{model[len(prefix):]}"
            break
    # Pattern 2: insert missing slash before trailing letter (EA-21D -> EA-21/D)
    m = _EA_QM_SLASH_PATTERN.match(model)
    if m:
        return f"{m.group(1)}-{m.group(2)}/{m.group(3)}"
    # Pattern 3: MSD model + variant letter -> family (MSD-200 E -> MSD-200)
    m = _MSD_FAMILY_PATTERN.match(model)
    if m:
        return m.group(1)
    return model


class Catalog:
    def __init__(self, config: ValveTypeConfig, file_path: Path, data_dir: Path | None = None):
        self.config = config
        self.file_path = file_path
        # Root used to resolve extra_row_sources (which may live in a different
        # subfolder than this catalog's file). Defaults to the file's parent.
        self.data_dir = data_dir if data_dir is not None else file_path.parent
        self.rows: list[dict[str, Any]] = []
        self._key_to_col = {k: idx for k, idx, _ in config.cascade}
        self._load()

    def _load(self) -> None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                wb = openpyxl.load_workbook(
                    self.file_path, data_only=True, keep_vba=False, read_only=True
                )
            except PermissionError:
                # File is locked (almost always Excel has it open). Copy to a
                # temp file and read from there so the app launches anyway.
                tmp_dir = Path(tempfile.gettempdir()) / "valve-selector-cache"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / self.file_path.name
                shutil.copyfile(self.file_path, tmp_path)
                print(
                    f"[valve-selector] {self.file_path.name} is locked "
                    f"(Excel has it open). Loading from temp copy."
                )
                wb = openpyxl.load_workbook(
                    tmp_path, data_only=True, keep_vba=False, read_only=True
                )

        sheet_names = [s for s in wb.sheetnames if self.config.sheet_marker in s]
        if not sheet_names:
            raise RuntimeError(
                f"No catalog sheets in {self.file_path.name} matching "
                f"marker {self.config.sheet_marker!r}."
            )

        max_col = max(
            max(idx for _, idx, _ in self.config.cascade),
            max(idx for idx, _ in self.config.detail_columns),
            self.config.primary_col,
            self.config.secondary_col or 0,
            self.config.bto_col if self.config.show_bto_fos else 0,
        )

        # Skip rows whose PRIMARY code column is empty. For ball/butterfly/
        # actuators primary_col is 1 (same as the old `not raw[0]` check); for
        # Pharma the real code is c2 (c1 is a Power-Query "Name" column), so this
        # drops the 728 blank-code spacer rows that would otherwise load as
        # ghost SKUs with no Bare Valve Code.
        key_idx = self.config.primary_col - 1
        for sn in sheet_names:
            ws = wb[sn]
            # Capture the primary-column header (row 1) so we can drop any
            # REPEATED header rows embedded mid-data. Some source workbooks
            # (e.g. Control Valve Dashboard, concatenated from several
            # Power-Query exports) carry the header line many times — at rows
            # 643/964/1285/… as well as row 1. Without this they load as ghost
            # rows whose every cell is the column label ("Bare Valve Code" =
            # "Bare Valve Code"), polluting every dropdown with a junk option.
            # Real codes never equal the header text, so this is safe for all
            # catalogs.
            header_key = None
            for r_i, raw in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True)):
                if not raw or key_idx >= len(raw):
                    continue
                key_val = _norm(raw[key_idx])
                if r_i == 0:  # header row
                    header_key = (
                        str(key_val).strip().lower()
                        if key_val not in (None, "") else None
                    )
                    continue
                if key_val in (None, ""):
                    continue
                if header_key is not None and str(key_val).strip().lower() == header_key:
                    continue  # repeated header row embedded in the data
                row = {f"c{i+1}": _norm(v) for i, v in enumerate(raw)}
                self.rows.append(row)

        if not self.rows:
            raise RuntimeError(
                f"{self.file_path.name} parsed but contained no rows under "
                f"marker {self.config.sheet_marker!r}."
            )

        for src in self.config.enrichment_sources:
            self._apply_enrichment(src)

        for src in self.config.extra_row_sources:
            self._apply_extra_rows(src)

    def _apply_extra_rows(self, src: ExtraRowSource) -> None:
        """Append rows from `src` whose key isn't already in self.rows, remapping
        source columns to catalog columns via src.column_map. Non-destructive."""
        try:
            src_path = find_catalog_file(self.data_dir, src.file_substring, src.path_contains)
        except FileNotFoundError:
            print(
                f"[valve-selector]   extra-rows skipped for {self.config.key}: "
                f"no file matching '*{src.file_substring}*'.",
                flush=True,
            )
            return

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                wb = openpyxl.load_workbook(
                    src_path, data_only=True, keep_vba=False, read_only=True
                )
            except PermissionError:
                tmp_dir = Path(tempfile.gettempdir()) / "valve-selector-cache"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / src_path.name
                shutil.copyfile(src_path, tmp_path)
                wb = openpyxl.load_workbook(
                    tmp_path, data_only=True, keep_vba=False, read_only=True
                )

        sheet_names = [s for s in wb.sheetnames if src.sheet_marker in s]
        if not sheet_names:
            wb.close()
            print(
                f"[valve-selector]   extra-rows skipped: no sheet matching "
                f"{src.sheet_marker!r} in {src_path.name}.",
                flush=True,
            )
            return

        pat = re.compile(src.key_pattern) if src.key_pattern else None
        master_key = f"c{src.master_key_col}"
        existing = {
            str(_norm(r.get(master_key))).strip()
            for r in self.rows if r.get(master_key) not in (None, "")
        }
        max_col = max(src.key_col, max(s for s, _ in src.column_map))
        added = 0
        for sn in sheet_names:
            for raw in wb[sn].iter_rows(min_row=2, max_col=max_col, values_only=True):
                if not raw or src.key_col - 1 >= len(raw):
                    continue
                key = _norm(raw[src.key_col - 1])
                if key in (None, ""):
                    continue
                key = str(key).strip()
                if pat and not pat.match(key):
                    continue          # header/garbage row
                if key in existing:
                    continue          # SKU already in the catalog
                row = {}
                for s_col, m_col in src.column_map:
                    row[f"c{m_col}"] = _norm(raw[s_col - 1]) if s_col - 1 < len(raw) else None
                self.rows.append(row)
                existing.add(key)
                added += 1
        wb.close()
        print(
            f"[valve-selector]   added {added} extra rows from {src_path.name}.",
            flush=True,
        )

    def _apply_enrichment(self, src: EnrichmentSource) -> None:
        """Pull extra columns from `src` and merge into self.rows by join key."""
        try:
            src_path = find_catalog_file(self.file_path.parent, src.file_substring)
        except FileNotFoundError:
            print(
                f"[valve-selector]   enrichment skipped for {self.config.key}: "
                f"no file matching '*{src.file_substring}*'.",
                flush=True,
            )
            return

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                wb = openpyxl.load_workbook(
                    src_path, data_only=True, keep_vba=False, read_only=True
                )
            except PermissionError:
                tmp_dir = Path(tempfile.gettempdir()) / "valve-selector-cache"
                tmp_dir.mkdir(exist_ok=True)
                tmp_path = tmp_dir / src_path.name
                shutil.copyfile(src_path, tmp_path)
                wb = openpyxl.load_workbook(
                    tmp_path, data_only=True, keep_vba=False, read_only=True
                )

        sheet_names = [s for s in wb.sheetnames if src.sheet_marker in s]
        if not sheet_names:
            print(
                f"[valve-selector]   enrichment skipped: no sheet matching "
                f"{src.sheet_marker!r} in {src_path.name}.",
                flush=True,
            )
            return

        max_col = max(src.source_key_col, max(s for s, _ in src.columns))
        lookup: dict[Any, dict[int, Any]] = {}
        for sn in sheet_names:
            ws = wb[sn]
            for raw in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                if not raw:
                    continue
                key = _norm(raw[src.source_key_col - 1])
                if key is None:
                    continue
                lookup[key] = {
                    dest: _norm(raw[s - 1]) for s, dest in src.columns
                }

        master_key = f"c{src.master_key_col}"
        hits = 0
        for row in self.rows:
            mk = row.get(master_key)
            if mk is None:
                continue
            extra = lookup.get(mk)
            if not extra:
                continue
            for dest, val in extra.items():
                if val is None:
                    continue
                row[f"c{dest}"] = val
            hits += 1
        print(
            f"[valve-selector]   enriched {hits} rows from {src_path.name}.",
            flush=True,
        )

    def _filter(self, picks: dict[str, Any]) -> list[dict[str, Any]]:
        if not picks:
            return self.rows
        out = []
        for row in self.rows:
            ok = True
            for key, val in picks.items():
                col = self._key_to_col.get(key)
                if col is None:
                    continue
                cell = row.get(f"c{col}")
                if val == CASCADE_NULL_OPTION:
                    if not _is_blank_cell(cell):   # "None" pick -> match blanks only
                        ok = False
                        break
                elif str(cell) != str(val):
                    ok = False
                    break
            if ok:
                out.append(row)
        return out

    def options(self, picks: dict[str, Any]) -> dict[str, list[str]]:
        matched = self._filter(picks)
        out: dict[str, list[str]] = {}
        for key, col, _ in self.config.cascade:
            if key in picks:
                continue
            raw = [row.get(f"c{col}") for row in matched]
            vals = {v for v in raw if not _is_blank_cell(v)}
            opts = sorted(vals, key=lambda v: (len(str(v)), str(v)))
            # Any cascade field with blank rows among matches offers a
            # "Not Applicable" option so blank-valued SKUs stay selectable
            # (instead of the picker force-matching the sole non-blank value).
            if any(_is_blank_cell(v) for v in raw):
                opts.append(CASCADE_NULL_OPTION)
            out[key] = opts
        return out

    def resolve(self, picks: dict[str, Any]) -> dict[str, Any] | None:
        matched = self._filter(picks)
        if not matched:
            return None
        row = matched[0]
        detail: dict[str, Any] = {
            "match_count": len(matched),
            "fields": [
                {"label": label, "value": row.get(f"c{col}")}
                for col, label in self.config.detail_columns
            ],
            "primary": row.get(f"c{self.config.primary_col}"),
            "secondary": (
                row.get(f"c{self.config.secondary_col}")
                if self.config.secondary_col else None
            ),
        }
        if self.config.show_bto_fos:
            bto = row.get(f"c{self.config.bto_col}")
            detail["bto"] = bto
            try:
                detail["fos"] = float(bto) * 1.5 if bto not in (None, "") else None
            except (TypeError, ValueError):
                detail["fos"] = None
        paired_list = []
        # Dedup rule (user's choice 2026-06-04):
        #   * DOUBLE-ACTING pressure slots (label has BOTH "Double Acting" and
        #     "@ … bar", e.g. Ball's DA @3.5/@4/@5.5) show every pressure even
        #     when the SAME model repeats — so a valve whose DA actuator is
        #     ACT-050D at all three pressures shows THREE chips.
        #   * EVERYTHING ELSE (Spring-Return at any pressure, Electric, Linear,
        #     and the generic "alternative" columns) dedupes by code: one chip
        #     per distinct (target_type, model). This keeps Spring-Return from
        #     repeating the same code across pressures, while still surfacing
        #     genuinely new alternatives (e.g. ACT-063SR07).
        seen_models = set()  # (target_type, model) shown so far
        seen_slots = set()   # (target_type, model, label) — exact DA pressure-slot twins
        # Per-series actuator labels: look up the row's controlling-field value
        # (e.g. Series) so 5016 shows "A Port Close"/"B Port Close".
        _ov_key = self.config.cascade_override_key or (
            self.config.cascade[0][0] if self.config.cascade else None
        )
        _ov_col = None
        if _ov_key:
            for _k, _col, _lab in self.config.cascade:
                if _k == _ov_key:
                    _ov_col = _col
                    break
        _row_series = row.get(f"c{_ov_col}") if _ov_col else None
        for p in self.config.paired_actuators:
            paired_val = row.get(f"c{p.model_col}")
            if paired_val in (None, ""):
                continue
            model = _normalize_paired_model(str(paired_val).strip())
            # `#N/A` is an Excel error string from broken VLOOKUP cells in the
            # source — treat as no recommendation (see 2026-05-27 report).
            if model in ("", "#N/A"):
                continue
            # Dual-use column guard: skip values that aren't actuator models
            # (e.g. control valve Fail-Safe cells holding a pressure or 0).
            if p.require_prefix and not model.startswith(p.require_prefix):
                continue
            target_type = p.resolve_target_type(model)
            if "Double Acting" in p.label and "@" in p.label:   # DA pressure slot — keep each pressure
                slot = (target_type, model, p.label)
                if slot in seen_slots:
                    continue
                seen_slots.add(slot)
            elif (target_type, model) in seen_models:            # everything else — one chip per code
                continue
            seen_models.add((target_type, model))
            paired_list.append({
                "model": model,
                "target_type": target_type,
                "target_field": p.target_field,
                "label": p.label_for(_row_series),
            })
        if paired_list:
            detail["paired_actuators"] = paired_list
        return detail

    def cascade(self) -> list[dict[str, str]]:
        return [{"key": k, "label": label} for k, _idx, label in self.config.cascade]


def find_catalog_file(
    data_dir: Path, file_substring: str, path_contains: str | None = None
) -> Path:
    """Pick the most recently modified .xlsx/.xlsm under `data_dir` (recursively)
    whose name contains `file_substring`. Walking recursively lets the catalog
    files live in nested subfolders (e.g. `data/Valve/Ball Valve Data Set/`).

    `path_contains` optionally restricts matches to files whose full path
    contains that fragment — used to disambiguate same-named files in different
    subfolders (e.g. the Pune vs Mumbai `Valve_Code_Selector_Dashboard.xlsx`)."""
    def _path_ok(p: Path) -> bool:
        if path_contains is None:
            return True
        return path_contains in str(p).replace("\\", "/")

    candidates = [
        p for p in data_dir.rglob("*.xls*")
        if file_substring in p.name and not p.name.startswith("~$") and _path_ok(p)
    ]
    if not candidates:
        scope = f" (path containing '{path_contains}')" if path_contains else ""
        raise FileNotFoundError(
            f"No file matching '*{file_substring}*'{scope} under {data_dir}."
        )
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def build_category_blocks(sections, planned_subgroups=None):
    """Group section summaries into category -> ordered subgroups for the picker.

    Each returned block: {"name": <category>, "subgroups": [
        {"name": <subgroup>, "members": [<section>...], "placeholder": <str>}]}.

    `planned_subgroups` (category -> [(subgroup_name, placeholder_text), ...])
    appends subgroups that have no loaded members yet, carrying placeholder text
    the template renders as a non-selectable row. A planned subgroup that already
    has loaded members is left untouched (real data wins). Planned entries for a
    category with no loaded sections are skipped — a placeholder needs a picker,
    and the picker only exists when the category has at least one family.
    """
    planned_subgroups = planned_subgroups or {}
    # Key each block by (group, subgroup), not subgroup name alone: two families
    # can share a blank subgroup under DIFFERENT parent groups (e.g. Ball/Control
    # sit directly under group "Pune" with no subgroup) and must NOT merge into
    # one block. The template still renders group headers + nested subgroups.
    categories: "OrderedDict[str, OrderedDict[tuple, list]]" = OrderedDict()
    for sec in sections:
        cat_key = sec["category"]
        if cat_key not in categories:
            categories[cat_key] = OrderedDict()
        grp = sec.get("group") or ""
        sub = sec.get("subgroup") or ""
        categories[cat_key].setdefault((grp, sub), []).append(sec)

    blocks = []
    for cat_name, groups in categories.items():
        block_subgroups = [
            {"name": sub, "members": members, "placeholder": "", "group": grp}
            for (grp, sub), members in groups.items()
        ]
        # Planned (empty) subgroups are superseded once ANY loaded family uses
        # that subgroup name, regardless of its parent group.
        present_names = {sub for (_grp, sub) in groups.keys()}
        for sub_name, placeholder in planned_subgroups.get(cat_name, []):
            if sub_name in present_names:
                continue  # already has real members
            block_subgroups.append(
                {"name": sub_name, "members": [], "placeholder": placeholder,
                 "group": ""}
            )
        blocks.append({"name": cat_name, "subgroups": block_subgroups})
    return blocks


def load_all(data_dir: Path) -> dict[str, Catalog]:
    out: dict[str, Catalog] = {}
    for cfg in VALVE_TYPES:
        try:
            path = find_catalog_file(data_dir, cfg.file_substring, cfg.path_contains)
        except FileNotFoundError:
            print(f"[valve-selector] Skipping {cfg.key}: no file matching '*{cfg.file_substring}*'.", flush=True)
            continue
        print(f"[valve-selector] Loading {cfg.key} catalog from {path.name}...", flush=True)
        out[cfg.key] = Catalog(cfg, path, data_dir)
        print(f"[valve-selector]   loaded {len(out[cfg.key].rows)} {cfg.key} rows.", flush=True)
    if not out:
        raise RuntimeError(f"No catalog files found in {data_dir}.")
    return out
