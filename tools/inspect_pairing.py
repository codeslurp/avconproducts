"""Read-only spot check: do ball-valve cols 49-57 contain real actuator codes?

Prints, for the first N ball-valve rows:
  - the valve's Bare Valve Code (col 1) and BTO (col 39)
  - each of the 9 pairing cell values (cols 49-57)
  - whether each pairing value matches a Code in any actuator catalog

Touches nothing. Just reads.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / "data"

BALL_FILE = next(DATA_DIR.glob("Ball Valve Data Sheet*"))
RP_FILE   = next(DATA_DIR.glob("Rack & Pinion Data Sheet*"))
SY_FILE   = next(DATA_DIR.glob("Scotch Yoke Actuator Data Sheet*"))
ER_FILE   = next(DATA_DIR.glob("Electrical Actuator Rotory Data Sheet*"))

BALL_MARKER = "BV NEW CODEING"
RP_MARKER   = "Rack & Pinion ACT"
SY_MARKER   = "Scotch Yoke Actuator SYA"
ER_MARKER   = "Electrical Actuator"

PAIRING_LABELS = [
    "DA Actuator 1", "DA Actuator 2", "DA Actuator 3",
    "SR-FC Actuator 1", "SR-FC Actuator 2", "SR-FC Actuator 3",
    "SR-FO Actuator 1", "SR-FO Actuator 2", "SR-FO Actuator 3",
]
PAIRING_COLS = list(range(49, 58))  # 49..57 inclusive (1-based)

ROWS_TO_INSPECT = 6  # first N data rows


def load_codes(path: Path, sheet_marker: str) -> set[str]:
    """Return set of Code values (col 1) across all matching sheets."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    out: set[str] = set()
    for sn in wb.sheetnames:
        if sheet_marker not in sn:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            code = row[0]
            if code is None:
                continue
            out.add(str(code).strip())
    return out


def main() -> int:
    print(f"Reading catalogs from {DATA_DIR}\n")
    rp_codes = load_codes(RP_FILE, RP_MARKER)
    sy_codes = load_codes(SY_FILE, SY_MARKER)
    er_codes = load_codes(ER_FILE, ER_MARKER)
    print(f"Loaded Code sets: {len(rp_codes)} R&P, {len(sy_codes)} Scotch Yoke, {len(er_codes)} Electrical Rotary\n")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(BALL_FILE, data_only=True, keep_vba=False, read_only=True)
    sheets = [s for s in wb.sheetnames if BALL_MARKER in s]
    print(f"Ball valve sheets with marker {BALL_MARKER!r}: {sheets}\n")

    # walk first sheet, first N rows
    ws = wb[sheets[0]]
    max_col = 57
    inspected = 0
    pairing_total = 0
    pairing_filled = 0
    pairing_matched = 0
    for raw in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        if not raw or not raw[0]:
            continue
        inspected += 1
        bare_code = raw[0]
        bto = raw[38] if len(raw) > 38 else None
        print(f"--- Row {inspected}: Bare Valve Code = {bare_code!r}  |  BTO col39 = {bto!r} ---")
        for label, col in zip(PAIRING_LABELS, PAIRING_COLS):
            val = raw[col - 1] if len(raw) >= col else None
            pairing_total += 1
            if val is None or (isinstance(val, str) and not val.strip()):
                status = "(empty)"
            else:
                pairing_filled += 1
                s = str(val).strip()
                if s in rp_codes:
                    status = "MATCH in Rack & Pinion"
                    pairing_matched += 1
                elif s in sy_codes:
                    status = "MATCH in Scotch Yoke"
                    pairing_matched += 1
                elif s in er_codes:
                    status = "MATCH in Electrical Rotary"
                    pairing_matched += 1
                else:
                    status = "NO MATCH (string not found in any actuator Code set)"
            print(f"  col {col:>2}  {label:<22} = {val!r:<60} -> {status}")
        print()
        if inspected >= ROWS_TO_INSPECT:
            break

    print(f"\nSummary across {inspected} valve rows ({pairing_total} pairing cells):")
    print(f"  filled  = {pairing_filled}/{pairing_total} ({pairing_filled / pairing_total:.0%})")
    if pairing_filled:
        print(f"  matched = {pairing_matched}/{pairing_filled} ({pairing_matched / pairing_filled:.0%} of filled cells)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
