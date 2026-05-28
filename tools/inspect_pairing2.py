"""Follow-up: 'ACT-050D' didn't match a Code exactly. Does it match a Model (col 5)?
Or appear as a substring of any actuator Code? Also: across the WHOLE ball-valve
catalog, what distinct values appear in cols 49-57?
"""
from __future__ import annotations

import warnings
from collections import Counter
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / "data"

BALL_FILE = next(DATA_DIR.glob("Ball Valve Data Sheet*"))
RP_FILE   = next(DATA_DIR.glob("Rack & Pinion Data Sheet*"))
SY_FILE   = next(DATA_DIR.glob("Scotch Yoke Actuator Data Sheet*"))
ER_FILE   = next(DATA_DIR.glob("Electrical Actuator Rotory Data Sheet*"))


def load_two_cols(path: Path, sheet_marker: str) -> tuple[set[str], set[str], list[str]]:
    """(code_set, model_set, sample_codes_for_display)."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    codes: set[str] = set()
    models: set[str] = set()
    sample: list[str] = []
    for sn in wb.sheetnames:
        if sheet_marker not in sn:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            code, _, _, _, model = row[:5]
            if code is not None:
                s = str(code).strip()
                codes.add(s)
                if len(sample) < 6:
                    sample.append(s)
            if model is not None:
                models.add(str(model).strip())
    return codes, models, sample


def collect_pairing_values(path: Path, marker: str) -> Counter:
    """Count distinct values across cols 49-57 in every ball-valve row."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    counts: Counter = Counter()
    for sn in wb.sheetnames:
        if marker not in sn:
            continue
        ws = wb[sn]
        for raw in ws.iter_rows(min_row=2, max_col=57, values_only=True):
            if not raw or not raw[0]:
                continue
            for col_idx in range(48, 57):  # 0-based 48..56 = 1-based 49..57
                v = raw[col_idx] if len(raw) > col_idx else None
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    counts[s] += 1
    return counts


print(f"Ball file: {BALL_FILE.name}\n")

rp_codes, rp_models, rp_sample = load_two_cols(RP_FILE, "Rack & Pinion ACT")
sy_codes, sy_models, sy_sample = load_two_cols(SY_FILE, "Scotch Yoke Actuator SYA")
er_codes, er_models, er_sample = load_two_cols(ER_FILE, "Electrical Actuator")

print(f"R&P: {len(rp_codes)} codes, {len(rp_models)} distinct models")
print(f"  sample codes: {rp_sample}")
print(f"  sample models: {sorted(rp_models)[:6]}")
print()
print(f"Scotch Yoke: {len(sy_codes)} codes, {len(sy_models)} distinct models")
print(f"  sample codes: {sy_sample}")
print(f"  sample models: {sorted(sy_models)[:6]}")
print()
print(f"Electrical Rotary: {len(er_codes)} codes, {len(er_models)} distinct models")
print(f"  sample codes: {er_sample}")
print(f"  sample models: {sorted(er_models)[:6]}")
print()

all_model_names = rp_models | sy_models | er_models
all_codes = rp_codes | sy_codes | er_codes

pairing_counts = collect_pairing_values(BALL_FILE, "BV NEW CODEING")
print(f"Across the entire ball-valve catalog, cols 49-57 contain "
      f"{sum(pairing_counts.values())} non-empty cells across "
      f"{len(pairing_counts)} distinct values.\n")

print("Top distinct pairing values found in ball-valve cols 49-57:")
for val, count in pairing_counts.most_common(20):
    in_code  = val in all_codes
    in_model = val in all_model_names
    substr_hit_code  = any(val in c for c in all_codes)  if not in_code  else False
    substr_hit_model = any(val in m for m in all_model_names) if not in_model else False
    flags = []
    if in_code:  flags.append("=Code")
    if in_model: flags.append("=Model")
    if substr_hit_code:  flags.append("substr-of-Code")
    if substr_hit_model: flags.append("substr-of-Model")
    if not flags: flags = ["NO MATCH anywhere"]
    print(f"  {count:>6}x  {val!r:<30}  ->  {', '.join(flags)}")
